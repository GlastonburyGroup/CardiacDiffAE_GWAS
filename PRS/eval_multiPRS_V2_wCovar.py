# %%
import warnings
try:
    from pandas.errors import SettingWithCopyWarning
    warnings.simplefilter(action='ignore', category=SettingWithCopyWarning)
except:
    pass
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
warnings.simplefilter(action='ignore', category=RuntimeWarning)

import os
import pandas as pd
from glob import glob
from tqdm import tqdm
import pickle
import sys
import itertools
import argparse

# sys.path.insert(0, os.getcwd())
sys.path.insert(0, "/group/glastonbury/soumick/MyCodes/GitLab/tricorder")

import numpy as np
from scipy import stats
from scipy.stats import norm
from statsmodels.stats.power import tt_ind_solve_power
from sklearn.metrics import roc_auc_score, f1_score
from sklearn.linear_model import LinearRegression

from skimage.filters import threshold_otsu

import pyreadr
from rds2py import read_rds

from openpyxl.styles import Font, Alignment, PatternFill

import matplotlib.pyplot as plt
import seaborn as sns

from lifelines import KaplanMeierFitter, NelsonAalenFitter, CoxPHFitter

from utils.stats.delong import delong_roc_test
from utils.stats.power import simulate_power_wilcoxon
from utils.python_utils import recursive_defaultdict, transform_text

from analyses.disease_mappings import *

import matplotlib.font_manager as fm
font_path = '/project/ukbblatent/soumick/fonts/Helvetica/Helvetica.ttf'  
font_prop = fm.FontProperties(fname=font_path, size=7)
fm.fontManager.addfont(font_path)
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = [font_prop.get_name()]
plt.rcParams['pdf.fonttype'] = 42 #Type 3 fonts (TrueType fonts), which are embedded in the PDF and kept as text rather than being converted to outlines.

# sns.set_style("whitegrid")

# for the prevalence plots, defination of the percentiles
percentiles = [5, 10, 20, 80, 90, 95]
x_labels = ['Bottom 5%', 'Bottom 10%', 'Bottom 20%', 'Top 20%', 'Top 10%', 'Top 5%']

# for the prevalence plots, defination of the colours and markers
# colours = ['salmon', 'turquoise', 'mediumpurple', 'sandybrown', 'lightgreen', 'palevioletred']
colours = ['#FF4D6FFF', '#579EA4FF', '#86AD34FF', '#5D7298FF', '#7E1A2FFF', '#C8350DFF']
markers = ['o', 's', 'D', '^', 'v', 'P']

# %% command-line arguments

def getARGSParser():
    parser = argparse.ArgumentParser(description='MultiPRS Script')
    # parser.add_argument('--output_root', type=str, help='Path to the multi-PRS analyses root directory', default="/group/glastonbury/soumick/PRS/LDPred2/F20208v3_DiffAE/nonDisc/newcovsets_V0v2/4paper_caucasian_king0p0625_grouped/panCohortV2_auto_lw_gw_10kIT_kingB4ldpred2")
    # parser.add_argument('--output_root', type=str, help='Path to the multi-PRS analyses root directory', default="/group/glastonbury/soumick/PRS/LDPred2/F20208v3_DiffAE/nonDisc/newcovsets_V0v2/4paper_caucasian_king0p0625_grouped/panCohortV2_NyDiag_auto_lw_gw_10kIT_kingB4ldpred2")
    parser.add_argument('--output_root', type=str, help='Path to the multi-PRS analyses root directory', default="/group/glastonbury/soumick/PRS/LDPred2/F20208v3_DiffAE/nonDisc/newcovsets_V0v2/4paper_caucasian_king0p0625_grouped/panCohortV2_1to10yProg_auto_lw_gw_10kIT_kingB4ldpred2")
    parser.add_argument('--disease_root', type=str, help='path to the root directory where the disease cohort fils are storred', default="/project/ukbblatent/clinicaldata/binary_disease_cohorts/F20208v3_nonDiscov/caucasian_king0p0625_grouped/newcovsets/V0v2")

    parser.add_argument('--rds_pres_prefix', type=str, help='Prefix (full path) to the PRS file', default="/group/glastonbury/soumick/PRS/LDPred2/F20208v3_DiffAE/nonDisc/run_ext_basic_king0p0625_lw_gw_indep_FiltMAF_")
    parser.add_argument('--rds_pres_suffix', type=str, help='Suffix after the pheno name in the RDS file name', default=".fullDS.auto.mod.LDPred2.rds")
    parser.add_argument('--rds_tag_prs', type=str, help='tag PRS present in the rds file name', default="auto.mod")
    parser.add_argument('--tag_data', type=str, help='tag PRS model', default="resNdata.basic")
    parser.add_argument('--tag_prs', type=str, help='tag PRS inside the rds file', default="pred_auto")

    parser.add_argument('--ext_covar', type=str, help='tag PRS inside the rds file', default="/group/glastonbury/soumick/PRS/inputs/F20208v3_DiffAE_select_latents_r80_discov_INF30/covars/nonDisc_caucasian_king0p0625_V0.tsv")
    parser.add_argument('--covar_cont_cols', type=str, help='Comma-separated list of continuous covariate column names (e.g., Age,PC1,PC2,PC3). Leave blank to use all columns starting with PC plus Age', default="Age,BMI")
    parser.add_argument('--covar_cat_cols', type=str, help='Comma-separated list of categorical covariate column names (e.g., Sex,AssessmentCentre). Leave blank for none', default="Sex,CAT_Smoking")
    parser.add_argument('--adjust_prs_for_covars', action=argparse.BooleanOptionalAction, default=True, help='Whether to adjust raw PRS for covariates (regress out covariate effects) in prevalence and disease plots')
    
    parser.add_argument('--reprocess_raw_diseases', action=argparse.BooleanOptionalAction, default=False, help='Whether to re-process the raw_diseaes, even if the file already exists in the current output_root')
    parser.add_argument('--raw_disease_path', type=str, help='Path to the raw/unprocessed disease file (not disease cohorts)', default="/project/ukbblatent/clinicaldata/merge_SR_HI_GP_v4_allUKB_&_HEALTHY.csv")
    parser.add_argument('--raw_baseline_path', type=str, help='Path to the raw/unprocessed baseline file (raw output of UKBPuller)', default="/project/ukbblatent/clinicaldata/v1.1.0_seventh_basket/baseline_MD_27_10_2023_13_10_05.tsv")
    parser.add_argument('--raw_centre_info_path', type=str, help='Path to the raw/unprocessed assessment centre info file (raw output of UKBPuller)', default="/project/ukbblatent/clinicaldata/v1.1.0_seventh_basket/assessmentCentre_82779_MD_13_06_2024_12_18_59.tsv")

    parser.add_argument('--drop_dis', type=str, help='Coma-seperated list of diseaes to drop', default="extended arrythmias atrial_ventricular,atrial fibrillation_flutter,arrythmias atrial_ventricular,mitral problems,extended miscellaneous,miscellaneous,aortic valve disorders,pericardial problem,cardiac arrest")
    parser.add_argument('--drop_comparisons', action=argparse.BooleanOptionalAction, default=True, help='Whether to drop comparisons if they are not marked as 1 in the list of comparisons-tags')
    
    parser.add_argument('--is_pancohort', action=argparse.BooleanOptionalAction, default=True, help='Whether to run/plot the pancohort probability predictions')
    parser.add_argument('--save_plots', action=argparse.BooleanOptionalAction, default=True, help='Whether to save the plots')
    parser.add_argument('--save_format', type=str, default="pdf", help='Whether to save the plots')
    parser.add_argument('--sex_stratified', action=argparse.BooleanOptionalAction, default=True, help='Whether to perform in sex-stratified manner as well')

    parser.add_argument('--obtain_summary', action=argparse.BooleanOptionalAction, default=False, help='Whether to obtain summary scores and perform statistical stats. If set to False, it will be presumed that they are already performed and the required files are present')
    parser.add_argument('--obtain_pairwise_improvements', action=argparse.BooleanOptionalAction, default=False, help='Whether to print the pairwise improvements')
    
    parser.add_argument('--split_top_dis_half', action=argparse.BooleanOptionalAction, default=False, help='Whether to split the diseases in half, as top and bottom diseases')
    parser.add_argument('--sort_mode_top_dis', type=str, default="0", help='0: score-based, 1: alphabetically')
    parser.add_argument('--n_top_dis', type=int, default=0, help='Number of top-AUC diseases to keep for plotting and export (0 = all diseases)')
    
    parser.add_argument('--plot_box_AUC', action=argparse.BooleanOptionalAction, default=False, help='Whether to plot the box plots for AUC')
    parser.add_argument('--plot_box_F1', action=argparse.BooleanOptionalAction, default=False, help='Whether to plot the box plots for F1')
    parser.add_argument('--plot_box_logOR', action=argparse.BooleanOptionalAction, default=False, help='Whether to plot the box plots for logOR [If pancohort, both pancohort and disease cohort will be plotted]')
    parser.add_argument('--plot_prevalence_prs_pancohort', action=argparse.BooleanOptionalAction, default=True, help='Whether to plot the prevalence vs PRS scores (best performing single PRS) plots for the pancohort')
    parser.add_argument('--plot_prevalence_prob_pancohort', action=argparse.BooleanOptionalAction, default=True, help='Whether to plot the prevalence vs probability plots for the pancohort')

    parser.add_argument('--plot_cum_disease_burden', action=argparse.BooleanOptionalAction, default=False, help='[Disease plot] Whether to plot the cumulative disease burden plots')
    parser.add_argument('--plot_cum_hazard', action=argparse.BooleanOptionalAction, default=False, help='[Disease plot] Whether to plot the cumulative hazard plots')
    parser.add_argument('--plot_KM_survival', action=argparse.BooleanOptionalAction, default=False, help='[Disease plot] Whether to plot the Kaplan-Meier survival plots')
    parser.add_argument('--plot_cox_ph', action=argparse.BooleanOptionalAction, default=False, help='[Disease plot] Whether to perform Cox Proportional Hazards regression')
    parser.add_argument('--dis_plots_mod', default="GLM,singlePRSCovar", help='Which predictive model to use for the disease plots. Leave it blank if not desired')
    parser.add_argument('--dis_plots_rawPRS', action=argparse.BooleanOptionalAction, default=True, help='[Disease plot] Whether to plot disease plots using raw PRS scores (Latent selected using the best performing single PRS for that disease)')
    parser.add_argument('--dis_plots_cutoff_date', default="2023-10-31", help='The cutoff date to be used (ideally, the download date) for the disease plots - to filter data with strange dates')
    parser.add_argument('--dis_plots_upto_Nyear', default=10, type=int, help='How many years to consider in the disease plots')    
    
    parser.add_argument('--save_raw', action=argparse.BooleanOptionalAction, default=True, help='[Disease and prevalence plots] Whether to store the raw values used to create the plots')
    
    parser.add_argument('--quick_run', action=argparse.BooleanOptionalAction, default=False, help='[Quick run mode] Enable quick run mode for targeted execution of specific plots/sex/diseases. When enabled, raw_store files are tagged to avoid overwrites.')
    parser.add_argument('--quick_run_sex', type=str, default="", help='[Quick run mode] Comma-separated list of sex groups to run (Both, Female, Male). Leave blank to run all three. Example: "Both" or "Female,Male"')
    parser.add_argument('--quick_run_diseases', type=str, default="", help='[Quick run mode] Comma-separated list of disease names to run. Leave blank to run all diseases. Example: "heart failure,diabetes"')

    parser.add_argument('--box_limits_AUC', type=str, default="0.55,0.80,0.05", help='Upper bound, lower bound, and distance for the y-axis of the AUC plots [Leave blank for using default]')
    parser.add_argument('--box_limits_F1', type=str, default="0.50,0.75,0.05", help='Upper bound, lower bound, and distance for the y-axis of the AUC plots [Leave blank for using default]')
    parser.add_argument('--box_limits_logOR', type=str, default="0.50,2.25,0.25", help='Upper bound, lower bound, and distance for the y-axis of the AUC plots [Leave blank for using default]')
    parser.add_argument('--box_limits_logOR_pancohort', type=str, default="0.50,2.00,0.25", help='Upper bound, lower bound, and distance for the y-axis of the AUC plots [Leave blank for using default]')
    
    parser.add_argument('--colour_box', type=str, default="#A4203D,#D87412,#5A6F94", help='List of colours to be used with the box plots [Leave blank for using default]')
    parser.add_argument('--colour_transparency_box', type=float, default=1, help='The colour_transparency percentage between 0 and 1 for the colours of the box plots [Set it to 0 or 1 to ignore]')
    parser.add_argument('--colour_prevalence', type=str, default="#FD4C6E,#80A531,#559AA0", help='List of colours to be used with the prevalence plots [Leave blank for using default]')
    parser.add_argument('--colour_displots', type=str, default="#C7350D,#F6BD01,#81B18C", help='List of colours to be used with the disease plots [Leave blank for using default]')

    return parser

# %% Decorators
def plots_and_results_by_sex(func):
    def wrapper(*args, **kwargs):
        # Check if sex_groups is specified (for quick run mode)
        if 'sex_groups' in kwargs:
            sex = kwargs['sex_groups']
            del kwargs['sex_groups']  # Remove it before passing to the function
        else:
            sex = ['Both', 'Female', 'Male']
        
        results = []
        for s in sex:
            new_kwargs = kwargs.copy()
            if s == 'Female' and 'IIDs_female' in kwargs:
                new_kwargs['IIDs'] = kwargs['IIDs_female']
            elif s == 'Male' and 'IIDs_male' in kwargs:
                new_kwargs['IIDs'] = kwargs['IIDs_male']
            elif 'IIDs_female' in kwargs and 'IIDs_female' in kwargs:
                new_kwargs['IIDs'] = kwargs['IIDs_female'] + kwargs['IIDs_male']
            if 'IIDs' in new_kwargs:
                del new_kwargs['IIDs_female']
                del new_kwargs['IIDs_male']
            new_kwargs['sex'] = s
            result = func(*args, **new_kwargs)
            results.append(result)
        return results
    return wrapper

# %% Support functions for the processing of the raw results and compute the stats

def get_stats(score_m0, score_m1):
    try:
        t_stat, p_value_t_test = stats.ttest_rel(score_m0, score_m1)
        print("Paired t-test - p-value:", p_value_t_test, f"({'Sig' if p_value_t_test < 0.05 else 'Not Sig'})")

        effect_size = (np.mean(score_m0-score_m1) / np.std(score_m0-score_m1))

        power = tt_ind_solve_power(effect_size=effect_size, nobs1=5, alpha=0.05, ratio=1, alternative='two-sided')
        print("Statistical Power (Paired t-test):", power)

        stat, p_value_wilcoxon = stats.wilcoxon(score_m0, score_m1)
        print("Wilcoxon Signed-Rank Test - p-value:", p_value_wilcoxon, f"({'Sig' if p_value_wilcoxon < 0.05 else 'Not Sig'})")
        
        estimated_power = simulate_power_wilcoxon(len(score_m0), effect_size)
        print("Estimated Power:", estimated_power)

        return p_value_t_test, power, p_value_wilcoxon, estimated_power
    except:
        return np.nan, np.nan, np.nan, np.nan

def get_sig(method0, method1, comboDF=None, pancohort=False):
    if type(method0) == str:
        method0 = comboDF[comboDF['model'] == method0]
    if type(method1) == str:
        method1 = comboDF[comboDF['model'] == method1]

    method0 = method0.sort_values('fold')
    method1 = method1.sort_values('fold')

    print("AUC:")
    p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC = get_stats(method0['AUC_test'].values, method1['AUC_test'].values)

    print("logOR:")
    p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR = get_stats(method0['logOR_test'].values, method1['logOR_test'].values)

    if pancohort:
        print("logOR (pancohort):")
        p_value_t_test_logOR_pancohort, power_t_test_logOR_pancohort, p_value_wilcoxon_logOR_pancohort, power_wilcoxon_logOR_pancohort = get_stats(method0['logOR_test_pancohort'].values, method1['logOR_test_pancohort'].values)

        return p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC, p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR, p_value_t_test_logOR_pancohort, power_t_test_logOR_pancohort, p_value_wilcoxon_logOR_pancohort, power_wilcoxon_logOR_pancohort
    else:
        return p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC, p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR

def getSummaryDF(df, pancohort=False):
    cols = ['AUC_test', 'logOR_test', 'logOR_test_pancohort', 'F1_test'] if pancohort else ['AUC_test', 'logOR_test', 'F1_test']
    group_cols = ['tag', 'Sex'] if "Sex" in df else ['tag']
    median_df = df.groupby(group_cols)[cols].median()   
    iqr_df = (df.groupby(group_cols)[cols].quantile(0.75) - df.groupby(group_cols)[cols].quantile(0.25))
    merged_df = median_df.join(iqr_df, lsuffix='_median', rsuffix='_iqr')
    merged_df['AUC_test'] = merged_df.apply(lambda row: f"{row['AUC_test_median']:.4f} ± {row['AUC_test_iqr']:.4f}", axis=1)
    merged_df['logOR_test'] = merged_df.apply(lambda row: f"{row['logOR_test_median']:.4f} ± {row['logOR_test_iqr']:.4f}", axis=1)
    merged_df['F1_test'] = merged_df.apply(lambda row: f"{row['F1_test_median']:.4f} ± {row['F1_test_iqr']:.4f}", axis=1)
    if pancohort:
        merged_df['logOR_test_pancohort'] = merged_df.apply(lambda row: f"{row['logOR_test_pancohort_median']:.4f} ± {row['logOR_test_pancohort_iqr']:.4f}", axis=1)
    return merged_df[cols].reset_index()

# %% Support functions to creating summary tables

def extract_median_iqr(value):
    median, iqr = value.split(' ± ')
    return float(median), float(iqr)

def determine_winner(row, col0, col1):
    median0, iqr0 = extract_median_iqr(row[col0])
    median1, iqr1 = extract_median_iqr(row[col1])    
    if median0 > median1:
        return row['Method0']
    elif median0 < median1:
        return row['Method1']
    else:
        if iqr0 == iqr1:
            return 'Tie'
        return row['Method0'] if iqr0 < iqr1 else row['Method1']

def combineDFs(statsig, sumres):
    merged_df = statsig.merge(sumres, left_on=['Method0', 'Sex', 'Disease'], right_on=['tag', 'Sex', 'Disease'], suffixes=('', '_Method0'))
    merged_df = merged_df.merge(sumres, left_on=['Method1', 'Sex', 'Disease'], right_on=['tag', 'Sex', 'Disease'], suffixes=('_Method0', '_Method1'))

    combined_df = merged_df[['Disease', 'Sex', 'Comparison', 'Method0', 'Method1', 'p_value_DeLong_AUC', 
                            'AUC_test_Method0', 'logOR_test_Method0', 'logOR_test_pancohort_Method0', 'F1_test_Method0', 
                            'AUC_test_Method1', 'logOR_test_Method1', 'logOR_test_pancohort_Method1', 'F1_test_Method1', ]]

    combined_df['Significant'] = combined_df['p_value_DeLong_AUC'] < 0.05
    combined_df['AUCWinner'] = combined_df.apply(determine_winner, axis=1, args=('AUC_test_Method0', 'AUC_test_Method1'))
    combined_df['logORWinner'] = combined_df.apply(determine_winner, axis=1, args=('logOR_test_Method0', 'logOR_test_Method1'))
    combined_df['logORPancohortWinner'] = combined_df.apply(determine_winner, axis=1, args=('logOR_test_pancohort_Method0', 'logOR_test_pancohort_Method1'))
    combined_df['F1Winner'] = combined_df.apply(determine_winner, axis=1, args=('F1_test_Method0', 'F1_test_Method1'))

    return combined_df

def summary_table(combined_df, ignore_methods=[], diseases=[]):
    def determine_result(row, col):
        winner = row[col]
        if winner == 'Tie':
            return 'Tie'
        if row['Significant']:
            if winner == row['Method0']:
                return f'Significantly Better for {row["Method0"]}'
            else:
                return f'Significantly Better for {row["Method1"]}'
        else:
            if winner == row['Method0']:
                return f'Insignificantly Better for {row["Method0"]}'
            else:
                return f'Insignificantly Better for {row["Method1"]}'

    combined_df['AUCResult'] = combined_df.apply(determine_result, axis=1, args=('AUCWinner',))
    combined_df['logORResult'] = combined_df.apply(determine_result, axis=1, args=('logORWinner',))
    combined_df['logORPancohortResult'] = combined_df.apply(determine_result, axis=1, args=('logORPancohortWinner',))
    combined_df['F1Result'] = combined_df.apply(determine_result, axis=1, args=('F1Winner',))

    if bool(diseases):
        combined_df = combined_df[combined_df['Disease'].isin(diseases)]

    summary = []

    for sex in combined_df['Sex'].unique():
        for comparison in combined_df['Comparison'].unique():
            sub_df = combined_df[(combined_df['Sex'] == sex) & (combined_df['Comparison'] == comparison)]
            auc_counts = sub_df['AUCResult'].value_counts().to_dict()
            logor_counts = sub_df['logORResult'].value_counts().to_dict()
            logor_pancohort_counts = sub_df['logORPancohortResult'].value_counts().to_dict()
            f1_counts = sub_df['F1Result'].value_counts().to_dict()

            if sub_df['Method0'].iloc[0] in ignore_methods or sub_df['Method1'].iloc[0] in ignore_methods:
                continue
            
            summary.append({
                'Sex': sex,
                'Comparison': comparison,
                'Method0': sub_df['Method0'].iloc[0],
                'Method1': sub_df['Method1'].iloc[0],
                'AUC_Significantly_Better_for_Method0': auc_counts.get(f'Significantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'AUC_Significantly_Better_for_Method1': auc_counts.get(f'Significantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'AUC_Insignificantly_Better_for_Method0': auc_counts.get(f'Insignificantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'AUC_Insignificantly_Better_for_Method1': auc_counts.get(f'Insignificantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'logOR_Significantly_Better_for_Method0': logor_counts.get(f'Significantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'logOR_Significantly_Better_for_Method1': logor_counts.get(f'Significantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'logOR_Insignificantly_Better_for_Method0': logor_counts.get(f'Insignificantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'logOR_Insignificantly_Better_for_Method1': logor_counts.get(f'Insignificantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'logOR_Pancohort_Significantly_Better_for_Method0': logor_pancohort_counts.get(f'Significantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'logOR_Pancohort_Significantly_Better_for_Method1': logor_pancohort_counts.get(f'Significantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'logOR_Pancohort_Insignificantly_Better_for_Method0': logor_pancohort_counts.get(f'Insignificantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'logOR_Pancohort_Insignificantly_Better_for_Method1': logor_pancohort_counts.get(f'Insignificantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'F1_Significantly_Better_for_Method0': f1_counts.get(f'Significantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'F1_Significantly_Better_for_Method1': f1_counts.get(f'Significantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'F1_Insignificantly_Better_for_Method0': f1_counts.get(f'Insignificantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'F1_Insignificantly_Better_for_Method1': f1_counts.get(f'Insignificantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'Number_of_Diseases': len(sub_df)
            })

    summary_df = pd.DataFrame(summary)
    return summary_df

def create_excel_summary(summary_df, metric_col, output_path):
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    for sex in summary_df['Sex'].unique():
        filtered_df = summary_df[summary_df['Sex'] == sex]
        
        summary_data = []
        for _, row in filtered_df.iterrows():
            summary_data.append([
                row['Method0'],
                row['Method1'],
                row[f'{metric_col}_Significantly_Better_for_Method0'],
                row[f'{metric_col}_Insignificantly_Better_for_Method0'],
                row[f'{metric_col}_Significantly_Better_for_Method1'],
                row[f'{metric_col}_Insignificantly_Better_for_Method1']
            ])
        
        summary_formatted_df = pd.DataFrame(summary_data, columns=[
            'Method0', 'Method1', 'Sig_Better_Method0', 'Insig_Better_Method0', 'Sig_Better_Method1', 'Insig_Better_Method1'
        ])
        summary_formatted_df.to_excel(writer, index=False, sheet_name=sex, startrow=2, header=False)

        workbook = writer.book
        worksheet = writer.sheets[sex]

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        worksheet.merge_cells('A1:B1')
        worksheet['A1'] = 'Methods'
        worksheet['A1'].font = header_font
        worksheet['A1'].alignment = header_alignment
        worksheet['A1'].fill = header_fill

        worksheet.merge_cells('C1:D1')
        worksheet['C1'] = 'better Method0'
        worksheet['C1'].font = header_font
        worksheet['C1'].alignment = header_alignment
        worksheet['C1'].fill = header_fill

        worksheet.merge_cells('E1:F1')
        worksheet['E1'] = 'better Method1'
        worksheet['E1'].font = header_font
        worksheet['E1'].alignment = header_alignment
        worksheet['E1'].fill = header_fill

        sub_headers = ['Method0', 'Method1', 'Sig', 'Insig', 'Sig', 'Insig']
        for col_num, sub_header in enumerate(sub_headers, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.value = sub_header
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

    writer.save()

@plots_and_results_by_sex
def compareSummaryPairs(df, all_combinations, disease, sex, pancohort=False):
    collect = []
    for combination in all_combinations:
        comparison1, comparison2 = combination
        print(f"(Sex: {sex}) {comparison1[2]} [vs] {comparison2[2]}-----")

        try:
            df_sex = df[df.Sex == sex]
            df_method1 = df_sex[df_sex.apply(lambda row: (row['method'], row['res_type']) == (comparison1[0], comparison1[1]), axis=1)]
            df_method2 = df_sex[df_sex.apply(lambda row: (row['method'], row['res_type']) == (comparison2[0], comparison2[1]), axis=1)]
            
            if pancohort:
                p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC, p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR, p_value_t_test_logOR_pancohort, power_t_test_logOR_pancohort, p_value_wilcoxon_logOR_pancohort, power_wilcoxon_logOR_pancohort = get_sig(df_method1, df_method2, pancohort=True)
            else:
                p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC, p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR = get_sig(df_method1, df_method2)

            res = {
                    "Disease": disease,
                    "Sex": sex,
                    "Comparison": f"{comparison1[2]} [vs] {comparison2[2]}",
                    "Method0": comparison1[2],
                    "Method1": comparison2[2],
                    "p_value_t_test_AUC": p_value_t_test_AUC,
                    "power_t_test_AUC": power_t_test_AUC,
                    "p_value_wilcoxon_AUC": p_value_wilcoxon_AUC,
                    "power_wilcoxon_AUC": power_wilcoxon_AUC,
                    "p_value_t_test_logOR": p_value_t_test_logOR,
                    "power_t_test_logOR": power_t_test_logOR,
                    "p_value_wilcoxon_logOR": p_value_wilcoxon_logOR,
                    "power_wilcoxon_logOR": power_wilcoxon_logOR
                }
            if pancohort:
                res.update({
                    "p_value_t_test_logOR_pancohort": p_value_t_test_logOR_pancohort,
                    "power_t_test_logOR_pancohort": power_t_test_logOR_pancohort,
                    "p_value_wilcoxon_logOR_pancohort": p_value_wilcoxon_logOR_pancohort,
                    "power_wilcoxon_logOR_pancohort": power_wilcoxon_logOR_pancohort
                })
            collect.append(res)
        except:
            pass
        
    return pd.DataFrame(collect)

def getIndividualProbs(res_method, res_type, IIDs, get_pancohort=False, resDF=None):
    if "single" in res_type.lower():
        if resDF is not None:
            latent = resDF.iloc[0].Latent
            res_method = res_method.loc[latent]
        else:
            res_method = res_method.iloc[res_method['AUC_test'].argmax()]
    res = pd.DataFrame(res_method['pred_probs_test_pancohort']) if get_pancohort else pd.DataFrame(res_method['pred_probs_test'])
    res = res[res.index.isin(IIDs)]
    return res, res_method

@plots_and_results_by_sex
def compareAUCPairs(res, all_combinations, disease, diseaseDF, IIDs, sex, resDF):
    collect = []
    for combination in all_combinations:
        comparison1, comparison2 = combination
        print(f"(Sex: {sex}) {comparison1[2]} [vs] {comparison2[2]} -----")
        
        try:
            collect_fold = []
            for fold in res.keys():
                pred_prob_m0, _ = getIndividualProbs(res[fold][comparison1[0]][comparison1[1]], comparison1[1], IIDs=IIDs, resDF=resDF[(resDF.tag == comparison1[2]) & (resDF.Sex == sex) & (resDF.fold == fold)])
                pred_prob_m1, _ = getIndividualProbs(res[fold][comparison2[0]][comparison2[1]], comparison2[1], IIDs=IIDs, resDF=resDF[(resDF.tag == comparison2[2]) & (resDF.Sex == sex) & (resDF.fold == fold)])
                pred_prob = pred_prob_m0.join(pred_prob_m1, lsuffix="_m0", rsuffix="_m1")
                combined = pred_prob.join(diseaseDF[['BinCAT_Disease']])
                delongP = delong_roc_test(combined['BinCAT_Disease'], combined['predicted_m0'], combined['predicted_m1'])
                print(f"DeLong's p-value [{fold}]:", delongP, f"({'Sig' if delongP < 0.05 else 'Not Sig'})")
                collect_fold.append(combined)
            collect_fold = pd.concat(collect_fold)
            delongP = delong_roc_test(collect_fold['BinCAT_Disease'], collect_fold['predicted_m0'], collect_fold['predicted_m1'])
            print(f"DeLong's p-value [Across Folds]:", delongP, f"({'Sig' if delongP < 0.05 else 'Not Sig'})")

            collect.append(
                {
                    "Disease": disease,
                    "Sex": sex,
                    "Comparison": f"{comparison1[2]} [vs] {comparison2[2]}",
                    "Method0": comparison1[2],
                    "Method1": comparison2[2],
                    "p_value_DeLong_AUC": delongP
                }
            )
        except:
            pass
    return pd.DataFrame(collect)


# %% Support functions related to the plotting

def prevalence_and_ci(pred_probs, true_labels, percentile, confidence=0.95):
    if percentile < 50:
        group = pred_probs <= np.percentile(pred_probs, percentile)
    else:
        group = pred_probs >= np.percentile(pred_probs, percentile)

    prevalence = true_labels[group].mean()
    n = len(true_labels[group])
    se = np.sqrt(prevalence * (1 - prevalence) / n)  
    ci_half_width = se * norm.ppf((1 + confidence) / 2)  
    return prevalence, prevalence - ci_half_width, prevalence + ci_half_width, n

def z_test(p_hat, p_0, n):
    se = np.sqrt(p_0 * (1 - p_0) / n)  
    z = (p_hat - p_0) / se  
    p_value = 2 * (1 - norm.cdf(abs(z)))  # Two-tailed test
    return p_value

def is_significant(p_hat, p_0, n, alpha=0.05):
    p_value = z_test(p_hat, p_0, n)
    return p_value < alpha

def generate_sequence_plotdelta(N):
    sequence = []
    if N == 1:
        sequence.append(0)
    else:
        for i in range(-(N//2), N//2 + 1):
            sequence.append(i * 0.1)
        if N % 2 == 0:
            sequence.remove(0)
    return sequence

def plot_prevalence(df, res_keys=['Pred_Covariates', 'Pred_PRS'], res_labels=['Covariates', 'PRS'], res_colours=['orange', 'blue'], res_fmts=['o', 's'], col_disease='BinCAT_Disease', plot_title='prova!', individual_prev_line=False, save_path="", flip_PRS_sign=False, save_raw=False):
    #supply Pred_Covariates, Pred_PRS and BinCAT_Disease as columns of the DataFrame 

    if save_raw:
        prevalence_data = {'Female': [], 'Both': [], 'Male': []}
        ci_lower_data = {'Female': [], 'Both': [], 'Male': []}
        ci_upper_data = {'Female': [], 'Both': [], 'Male': []}
        n_data = {'Female': [], 'Both': [], 'Male': []}
        n_total = {}
        percentile_labels = ['Bottom 5%', 'Bottom 10%', 'Bottom 20%', 'Top 20%', 'Top 10%', 'Top 5%']

    plotdelta = generate_sequence_plotdelta(len(res_keys))
    x_ticks = np.arange(len(x_labels))

    overall_prevalence = df[col_disease].mean()
    mid_percentile = x_ticks.mean()

    plt.figure(constrained_layout=True)
    fig, ax = plt.subplots(figsize=(10, 6))

    for i in range(len(res_keys)):
        df_curr = df[[res_keys[i], col_disease]].dropna() #this will handle the sex-stratified case
        results = [prevalence_and_ci(-df_curr[res_keys[i]] if flip_PRS_sign else df_curr[res_keys[i]], df_curr[col_disease], p) for p in percentiles]
        bottom, bottom_cis_lower, bottom_cis_upper, bottom_n = zip(*results[:3])
        top, top_cis_lower, top_cis_upper, top_n = zip(*results[3:])        

        if save_raw:
            prevalence_data[res_labels[i]].extend(bottom + top)
            ci_lower_data[res_labels[i]].extend(bottom_cis_lower + top_cis_lower)
            ci_upper_data[res_labels[i]].extend(bottom_cis_upper + top_cis_upper)
            n_data[res_labels[i]].extend(list(bottom_n) + list(top_n))
            n_total[res_labels[i]] = len(df_curr)
        
        ax.errorbar(np.arange(len(bottom)) + plotdelta[i], bottom, yerr=[np.subtract(bottom, bottom_cis_lower), np.subtract(bottom_cis_upper, bottom)], fmt=res_fmts[i], color=res_colours[i], label=res_labels[i], capsize=5)
        ax.errorbar(np.arange(len(top)) + 3 + plotdelta[i], top, yerr=[np.subtract(top, top_cis_lower), np.subtract(top_cis_upper, top)], fmt=res_fmts[i], color=res_colours[i], capsize=5)

        if individual_prev_line:
            ax.axhline(y=df_curr[col_disease].mean(), color=res_colours[i], linestyle='--')

    if not individual_prev_line:
        ax.axhline(y=overall_prevalence, color='gray', linestyle='--')
    ax.axvline(x=mid_percentile, color='gray', linestyle='dotted')

    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    ax.set_xticks(x_ticks)
    ax.set_xticklabels(x_labels, fontsize=13)
    ax.set_xlabel('Percentile', fontsize=15)
    ax.set_ylabel('Prevalence', fontsize=15)
    ax.set_title(plot_title, fontsize=16)
    ax.legend(fontsize=13, title_fontsize=15)

    ax.grid(False)

    if bool(save_path):
        plt.savefig(save_path, dpi=300)
    else:
        plt.show()

    if save_raw:
        return {"prevalence_data": prevalence_data, "ci_lower_data": ci_lower_data, "ci_upper_data": ci_upper_data, "n_data": n_data, "n_total": n_total, "percentile_labels": percentile_labels, "ci_method": "normal approximation of binomial proportion (95% CI)"}

@plots_and_results_by_sex
def plot_box(df, sex, x='Disease', y='AUC_test', y_tag='AUC', legends_tag="Method", hue='tag', hue_order=None, palette="Set2", colour_transparency=0.6, y_tick_limits="", save_path_noext="", save_format="pdf"):
    df_sex = df[df.Sex == sex]

    plt.figure(figsize=(14, 8), constrained_layout=True)

    ax = sns.boxplot(data=df_sex, x=x, y=y, hue=hue, hue_order=hue_order, palette=palette, saturation=1, linewidth=1)
    if colour_transparency not in [0,1]:
        for patch in ax.patches:
            r, g, b, a = patch.get_facecolor()
            patch.set_facecolor((r, g, b, colour_transparency))

    # ax.spines['left'].set_visible(False)
    # ax.spines['bottom'].set_visible(False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.title(f'{y_tag} by {x} and {legends_tag} (Sex: {sex})', fontsize=19)
    plt.xlabel(x, fontsize=18)
    plt.ylabel(y_tag, fontsize=18)
    plt.legend(title=legends_tag, fontsize=16, title_fontsize=18)
    plt.xticks(rotation=45, fontsize=16)
    plt.grid(False)
    # plt.grid(True, which='both', color='grey', linestyle='--', linewidth=0.5, alpha=0.4)

    if bool(y_tick_limits):
        y_tick_limits = [float(t) for t in y_tick_limits.split(",")]
        plt.ylim(y_tick_limits[0]-0.02, y_tick_limits[1]+0.02)
        plt.yticks(np.arange(y_tick_limits[0], y_tick_limits[1]+0.02, y_tick_limits[2]), fontsize=16)
    else:
        plt.yticks(fontsize=16)

    if bool(save_path_noext):
        plt.savefig(f"{save_path_noext}_{sex}.{save_format}", dpi=300)
    else:
        plt.show()

@plots_and_results_by_sex
def plot_prevalence_prob_pancohort(res, disease, diseaseDF, collect_df, sex, IIDs, comparisons, fullcohort=True, save_format="pdf"):
    best_fold = collect_df[(collect_df.Disease == disease) & (collect_df.res_type == "singlePRSCovar") & (collect_df.Sex == sex)].sort_values('AUC_test', ascending=False).iloc[0]

    probs = []
    tags = []
    for c in comparisons:
        try:
            resDF = collect_df[(collect_df.Disease == disease) & (collect_df.tag == c[2]) & (collect_df.Sex == sex) & (collect_df.fold == best_fold.fold)]
            pred_prob, _ = getIndividualProbs(res[best_fold.fold][c[0]][c[1]], c[1], IIDs=IIDs, resDF=resDF, get_pancohort=True)
            pred_prob.rename({'predicted': c[1]}, axis=1, inplace=True)
            probs.append(pred_prob)
            tags.append(c[2])
        except:
            pass
    probs = pd.concat(probs, axis=1)

    combined = probs.join(diseaseDF[['BinCAT_Disease']])

    if fullcohort:
        combined.fillna(0, inplace=True)
    else:
        combined
    combined["BinCAT_Disease"] = combined["BinCAT_Disease"].astype(int)

    res_cols = [c for c in combined.columns if c!="BinCAT_Disease"]

    if args.save_plots:
        covar_subdir = "covar_adjusted/" if args.adjust_prs_for_covars else ""
        pth = f"{args.output_root}/plots/{covar_subdir}prevalence_prob_pancohort/bestfold_{d.replace(' ', '-')}_{sex}.{save_format}"
    else:
        pth = ""
    return plot_prevalence(combined, res_keys=res_cols, res_labels=tags, res_colours=args.colour_prevalence[:len(tags)], res_fmts=markers[:len(tags)], col_disease='BinCAT_Disease', plot_title=f'Risk scores (Full cohort [Sex: {sex}], excluding discovery and PRS training sets): {d.replace("-", " ")}', save_path=pth)

@plots_and_results_by_sex
def plot_dis_plots(res, disease, diseaseDF, collect_df, sex, IIDs, args, covars=None, model_type="GLM", model="singlePRSCovar", PRS_col="predicted", ending_year = 10, use_raw_PRS=False, res_colours=['#FF4D6FFF', '#579EA4FF', '#86AD34FF'], save_format="pdf", save_raw=False):
    if use_raw_PRS:
        model_type = "GLM"
        model = "singlePRS"
        tag = "rawPRS_"
    else:
        tag = ""

    if save_raw:
        raw_store = {"Type": "bestfold" if tag=="" else "rawPRS", "Disease": disease, "Sex": sex}

    resDF = collect_df[(collect_df.Disease == disease) & (collect_df.method == model_type) & (collect_df.res_type == model) & (collect_df.Sex == sex)].sort_values('AUC_test', ascending=False)
    
    # Check if resDF is empty (no matching results found)
    if resDF.empty:
        print(f"\nWARNING: No results found for disease '{disease}' with model_type='{model_type}', res_type='{model}', Sex='{sex}'")
        # Show available combinations for this disease to help debug
        disease_results = collect_df[collect_df.Disease == disease]
        if not disease_results.empty:
            print(f"  Available combinations for '{disease}':")
            available = disease_results[['method', 'res_type', 'Sex']].drop_duplicates()
            for _, row in available.iterrows():
                print(f"    - method='{row['method']}', res_type='{row['res_type']}', Sex='{row['Sex']}'")
        else:
            print(f"  No results found for disease '{disease}' at all in collect_df")
        return None
    
    if use_raw_PRS:
        pred_prob = getBestPRS(rds_pres_prefix=args.rds_pres_prefix, rds_pres_suffix=args.rds_pres_suffix, rds_tag_prs=args.rds_tag_prs, tag_data=args.tag_data, tag_prs=args.tag_prs, IIDs=IIDs, resDF=resDF)
    else:
        pred_prob, _ = getIndividualProbs(res[resDF.iloc[0].fold][model_type][model], model, IIDs=IIDs, resDF=resDF, get_pancohort=True)
    
    # Merge covariates if provided for covariate adjustment
    if covars is not None:
        merge_cols = ['IID'] + (args.covar_cont_cols if args.covar_cont_cols else []) + (args.covar_cat_cols if args.covar_cat_cols else [])
        # Only include columns that exist in covars
        merge_cols = [c for c in merge_cols if c in covars.columns]
        pred_prob = pred_prob.merge(covars[merge_cols], left_index=True, right_on='IID', how='left').set_index('IID')
    
    prob_days = pred_prob.merge(diseaseDF[diseaseDF.summary == disease.lower()]['DiseaseAfter'], left_index=True, right_index=True)
    prob_control = pred_prob.merge(diseaseDF[diseaseDF.summary == "healthy"]['DiseaseAfter'], left_index=True, right_index=True)

    # Check if we have any disease subjects or controls after merging
    if prob_days.empty:
        print(f"\nWARNING: No disease subjects found for '{disease}' (Sex: {sex}) after merging with diseaseDF. Skipping disease plots.")
        return None

    if prob_control.empty:
        print(f"\nWARNING: No healthy controls found for '{disease}' (Sex: {sex}) after merging with diseaseDF. Skipping disease plots.")
        return None

    if args.plot_cum_disease_burden:
        # Adjust PRS for covariates if available
        if args.adjust_prs_for_covars and covars is not None and (args.covar_cont_cols or args.covar_cat_cols):
            # Use continuous covariates that exist in prob_days
            available_cont_covars = [c for c in args.covar_cont_cols if c in prob_days.columns] if args.covar_cont_cols else []
            available_cat_covars = [c for c in args.covar_cat_cols if c in prob_days.columns] if args.covar_cat_cols else []
            
            if available_cont_covars or available_cat_covars:
                X_parts = []
                if available_cont_covars:
                    X_cont = prob_days[available_cont_covars].fillna(prob_days[available_cont_covars].median())
                    X_parts.append(X_cont)
                if available_cat_covars:
                    X_cat = pd.get_dummies(prob_days[available_cat_covars], drop_first=True)
                    X_parts.append(X_cat)
                X_covars = pd.concat(X_parts, axis=1)
                lr = LinearRegression()
                lr.fit(X_covars, prob_days[PRS_col])
                prob_days['PRS_adjusted'] = prob_days[PRS_col] - lr.predict(X_covars)
                prob_days['PRSQuantile'] = pd.qcut(prob_days['PRS_adjusted'], q=3, labels=['Low risk PRS', 'Mid risk PRS', 'Top risk PRS'])
            else:
                prob_days['PRSQuantile'] = pd.qcut(prob_days[PRS_col], q=3, labels=['Low risk PRS', 'Mid risk PRS', 'Top risk PRS'])
        else:
            prob_days['PRSQuantile'] = pd.qcut(prob_days[PRS_col], q=3, labels=['Low risk PRS', 'Mid risk PRS', 'Top risk PRS'])
        prob_days['Years'] = prob_days['DiseaseAfter'] / 365.25
        time_periods = np.arange(1, prob_days['Years'].max()+1 if ending_year == -1 else ending_year)  
        quantiles = ['Top risk PRS', 'Mid risk PRS', 'Low risk PRS']

        plot_data = {quantile: [] for quantile in quantiles}

        for year in time_periods:
            for quantile in quantiles:
                count = prob_days[(prob_days['PRSQuantile'] == quantile) & (prob_days['Years'] <= year)].shape[0]
                plot_data[quantile].append(count)
        
        if save_raw:
            raw_store['cum_disease_burden'] = plot_data

        plt.figure(figsize=(6, 6), constrained_layout=True)
        for i, quantile in enumerate(quantiles):
            plt.plot(time_periods, plot_data[quantile], label=quantile, color=res_colours[i])

        plt.xlabel('Years', fontsize=15)
        plt.ylabel('Cumulative number of diagnoses', fontsize=15)
        plt.title(f'{model} for {disease} [Sex: {sex}]', fontsize=16)
        plt.legend(title='PRS Quantiles', fontsize=13, title_fontsize=15)
        plt.xticks(fontsize=13)
        plt.yticks(fontsize=13)
        plt.grid(False)

        ax = plt.gca()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        if args.save_plots:
            covar_subdir = "covar_adjusted/" if args.adjust_prs_for_covars else ""
            plt.savefig(f"{args.output_root}/plots/{covar_subdir}{tag}cumulative_disease_burden/bestfold_{disease.replace(' ', '-')}_{sex}.{save_format}", dpi=300)
        else:
            plt.show()

    if args.plot_cum_hazard or args.plot_KM_survival or args.plot_cox_ph:
        prob_days = prob_days[prob_days.DiseaseAfter <= 365*ending_year] #TODO: if we decide to keep it, make it a flag.
        
        # Check if filtering removed all disease subjects
        if prob_days.empty:
            print(f"\nWARNING: No disease subjects found for '{disease}' (Sex: {sex}) within {ending_year} year window. Skipping survival analysis plots.")
            return None
        
        # Include covariate columns if available
        if covars is not None and args.covar_cont_cols:
            # Use all covariates (continuous and categorical) that exist in prob_days
            all_covar_cont_cols = args.covar_cont_cols + args.covar_cat_cols
            available_covars = [c for c in all_covar_cont_cols if c in prob_days.columns]
            if available_covars:
                merge_cols = ["predicted", "DiseaseAfter"] + available_covars
                prob_incident_control = pd.concat([prob_days[merge_cols], prob_control])
            else:
                prob_incident_control = pd.concat([prob_days[["predicted", "DiseaseAfter"]], prob_control])
        else:
            prob_incident_control = pd.concat([prob_days[["predicted", "DiseaseAfter"]], prob_control])
        
        prob_incident_control['event_occurred'] = prob_incident_control['DiseaseAfter'].notna()
        prob_incident_control = prob_incident_control[prob_incident_control['DiseaseAfter'].gt(0) | prob_incident_control['DiseaseAfter'].isna()]
        prob_incident_control['DiseaseAfter'] = prob_incident_control['DiseaseAfter'].fillna(prob_incident_control['DiseaseAfter'].max())
        prob_incident_control['DiseaseAfter'] = prob_incident_control['DiseaseAfter'] / 365.25
        
        # Check if we have enough data after filtering
        if prob_incident_control.empty:
            print(f"\nWARNING: No subjects remaining for '{disease}' (Sex: {sex}) after filtering. Skipping survival analysis plots.")
            return None
        
        # Adjust PRS for covariates if available (use both continuous and categorical covariates for regression)
        if args.adjust_prs_for_covars and covars is not None and (args.covar_cont_cols or args.covar_cat_cols):
            available_cont_covars = [c for c in args.covar_cont_cols if c in prob_incident_control.columns] if args.covar_cont_cols else []
            available_cat_covars = [c for c in args.covar_cat_cols if c in prob_incident_control.columns] if args.covar_cat_cols else []
            
            if available_cont_covars or available_cat_covars:
                X_parts = []
                if available_cont_covars:
                    X_cont = prob_incident_control[available_cont_covars].fillna(prob_incident_control[available_cont_covars].median())
                    X_parts.append(X_cont)
                if available_cat_covars:
                    X_cat = pd.get_dummies(prob_incident_control[available_cat_covars], drop_first=True)
                    X_parts.append(X_cat)
                X_covars = pd.concat(X_parts, axis=1)
                lr = LinearRegression()
                lr.fit(X_covars, prob_incident_control[PRS_col])
                prob_incident_control['PRS_adjusted'] = prob_incident_control[PRS_col] - lr.predict(X_covars)
                prob_incident_control['PRSQuantile'] = pd.qcut(prob_incident_control['PRS_adjusted'], q=3, labels=['Low risk PRS', 'Mid risk PRS', 'Top risk PRS'])
            else:
                prob_incident_control['PRSQuantile'] = pd.qcut(prob_incident_control[PRS_col], q=3, labels=['Low risk PRS', 'Mid risk PRS', 'Top risk PRS'])
        else:
            prob_incident_control['PRSQuantile'] = pd.qcut(prob_incident_control[PRS_col], q=3, labels=['Low risk PRS', 'Mid risk PRS', 'Top risk PRS'])

        if save_raw:
            raw_store['n_incident_cases'] = int(prob_incident_control['event_occurred'].sum())
            raw_store['n_controls'] = int((~prob_incident_control['event_occurred']).sum())
            raw_store['n_total'] = len(prob_incident_control)
            raw_store['n_per_risk_group'] = {grp: int((prob_incident_control['PRSQuantile'] == grp).sum()) for grp in ['Low risk PRS', 'Mid risk PRS', 'Top risk PRS']}

        if args.plot_cum_hazard:
            naf = NelsonAalenFitter()
            if save_raw:
                raw_store['cum_hazard'] = {}

            plt.figure(figsize=(6, 6), constrained_layout=True)
            for i, risk_group in enumerate(['Top risk PRS', 'Mid risk PRS', 'Low risk PRS']):
                mask = prob_incident_control['PRSQuantile'] == risk_group
                naf.fit(durations=prob_incident_control['DiseaseAfter'][mask], event_observed=prob_incident_control['event_occurred'][mask], label=risk_group)
                if save_raw:
                    raw_store['cum_hazard'][risk_group] = {
                        'cumulative_hazard': naf.cumulative_hazard_,
                        'confidence_intervals': naf.confidence_interval_,
                        'n': int(mask.sum()),
                        'n_events': int(prob_incident_control['event_occurred'][mask].sum()),
                    }
                naf.plot_cumulative_hazard(color=res_colours[i])

            plt.title(f'{model} for {disease} [Sex: {sex}]', fontsize=16)
            plt.xlabel('Years after baseline', fontsize=15)
            plt.ylabel('Cumulative hazard', fontsize=15)
            plt.xlim(0, ending_year)  
            plt.legend(title='PRS risk group', fontsize=13, title_fontsize=15)
            plt.xticks(fontsize=13)
            plt.yticks(fontsize=13)
            plt.grid(False)

            ax = plt.gca()
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            if args.save_plots:
                covar_subdir = "covar_adjusted/" if args.adjust_prs_for_covars else ""
                plt.savefig(f"{args.output_root}/plots/{covar_subdir}{tag}cumulative_hazard/bestfold_{disease.replace(' ', '-')}_{sex}.{save_format}", dpi=300)
            else:
                plt.show()

        if args.plot_KM_survival:
            kmf = KaplanMeierFitter()
            if save_raw:
                raw_store['KM_survival'] = {}

            plt.figure(figsize=(6, 6), constrained_layout=True)
            for i, risk_group in enumerate(['Top risk PRS', 'Mid risk PRS', 'Low risk PRS']):
                mask = prob_incident_control['PRSQuantile'] == risk_group
                kmf.fit(durations=prob_incident_control['DiseaseAfter'][mask], event_observed=prob_incident_control['event_occurred'][mask], label=risk_group)
                if save_raw:
                    raw_store['KM_survival'][risk_group] = {
                        'survival_function': kmf.survival_function_,
                        'confidence_intervals': kmf.confidence_interval_,
                        'n': int(mask.sum()),
                        'n_events': int(prob_incident_control['event_occurred'][mask].sum()),
                    }
                kmf.plot(color=res_colours[i])

            plt.title(f'{model} for {disease} [Sex: {sex}]', fontsize=16)
            plt.xlabel('Years after baseline', fontsize=15)
            plt.ylabel('Survival probability', fontsize=15)
            plt.xlim(0, ending_year)  
            plt.legend(title='PRS risk group', fontsize=14, title_fontsize=15)
            plt.xticks(fontsize=13)
            plt.yticks(fontsize=13)
            plt.grid(False)

            ax = plt.gca()
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            if args.save_plots:
                covar_subdir = "covar_adjusted/" if args.adjust_prs_for_covars else ""
                plt.savefig(f"{args.output_root}/plots/{covar_subdir}{tag}KM_survival/bestfold_{disease.replace(' ', '-')}_{sex}.{save_format}", dpi=300)
            else:
                plt.show()

        if args.plot_cox_ph:
            cph = CoxPHFitter()
            if save_raw:
                raw_store['cox_ph'] = {}

            # Approach 1: Continuous PRS Analysis
            print(f"\n{'='*80}")
            print(f"Cox Proportional Hazards Analysis - Continuous PRS")
            print(f"Disease: {disease}, Sex: {sex}, Model: {model}")
            print(f"{'='*80}")
            
            # Build continuous model dataframe
            continuous_cols = ['DiseaseAfter', 'event_occurred']
            
            # Use adjusted or raw PRS based on covariate adjustment flag
            if args.adjust_prs_for_covars and 'PRS_adjusted' in prob_incident_control.columns:
                prs_continuous = prob_incident_control['PRS_adjusted'].copy()
            else:
                prs_continuous = prob_incident_control[PRS_col].copy()
            
            # Z-score standardise the PRS for "per SD increase" interpretation
            prs_z_scored = (prs_continuous - prs_continuous.mean()) / prs_continuous.std()
            
            cox_continuous_df = prob_incident_control[continuous_cols].copy()
            cox_continuous_df['PRS_z_scored'] = prs_z_scored
            
            # Add covariates if available and not pre-adjusted
            if covars is not None and not args.adjust_prs_for_covars:
                available_cont_covars = [c for c in args.covar_cont_cols if c in prob_incident_control.columns] if args.covar_cont_cols else []
                available_cat_covars = [c for c in args.covar_cat_cols if c in prob_incident_control.columns] if args.covar_cat_cols else []
                
                if available_cont_covars:
                    for col in available_cont_covars:
                        cox_continuous_df[col] = prob_incident_control[col].fillna(prob_incident_control[col].median())
                
                if available_cat_covars:
                    cat_dummies = pd.get_dummies(prob_incident_control[available_cat_covars], drop_first=True, prefix=available_cat_covars)
                    cox_continuous_df = pd.concat([cox_continuous_df, cat_dummies], axis=1)
            
            # Fit continuous Cox model
            cph.fit(cox_continuous_df, duration_col='DiseaseAfter', event_col='event_occurred')
            
            # Print summary statistics
            print("\nCox Model Summary (Continuous PRS):")
            print(cph.summary[['coef', 'exp(coef)', 'p', 'z']])
            print(f"\nConcordance Index: {cph.concordance_index_:.4f}")
            
            # Test proportional hazards assumption
            print("\nProportional Hazards Assumption Test:")
            try:
                ph_assumption = cph.check_assumptions(cox_continuous_df, p_value_threshold=0.05, show_plots=False)
                print("Proportional hazards assumption check completed.")
            except Exception as e:
                print(f"Warning: Could not test proportional hazards assumption: {e}")
            
            # Store results if save_raw
            if save_raw:
                raw_store['cox_ph']['continuous'] = {
                    'hazard_ratios': cph.hazard_ratios_.to_dict(),
                    'confidence_intervals': cph.confidence_intervals_.to_dict(),
                    'summary': cph.summary.to_dict(),
                    'concordance_index': cph.concordance_index_,
                    'aic_partial': cph.AIC_partial_,
                    'log_likelihood': cph.log_likelihood_
                }

            # Approach 2: Categorical Risk Groups Analysis (All Pairwise Comparisons)
            print(f"\n{'='*80}")
            print(f"Cox Proportional Hazards Analysis - Categorical Risk Groups")
            print(f"Disease: {disease}, Sex: {sex}, Model: {model}")
            print(f"{'='*80}")
            
            # Base categorical dataframe with covariates
            categorical_cols = ['DiseaseAfter', 'event_occurred', 'PRSQuantile']
            cox_categorical_base = prob_incident_control[categorical_cols].copy()
            
            # Add covariates if available and not pre-adjusted
            if covars is not None and not args.adjust_prs_for_covars:
                available_cont_covars = [c for c in args.covar_cont_cols if c in prob_incident_control.columns] if args.covar_cont_cols else []
                available_cat_covars = [c for c in args.covar_cat_cols if c in prob_incident_control.columns] if args.covar_cat_cols else []
                
                if available_cont_covars:
                    for col in available_cont_covars:
                        cox_categorical_base[col] = prob_incident_control[col].fillna(prob_incident_control[col].median())
                
                if available_cat_covars:
                    cat_dummies = pd.get_dummies(prob_incident_control[available_cat_covars], drop_first=True, prefix=available_cat_covars)
                    cox_categorical_base = pd.concat([cox_categorical_base, cat_dummies], axis=1)
            
            # Store all pairwise comparisons
            if save_raw:
                raw_store['cox_ph']['categorical'] = {
                    'pairwise_comparisons': {},
                    'concordance_index': None
                }
            
            # Fit models with each risk group as reference to get all pairwise comparisons
            risk_groups = ['Mid risk PRS', 'Low risk PRS', 'Top risk PRS']
            
            for ref_group in risk_groups:
                print(f"\n--- Reference Group: {ref_group} ---")
                
                # Create dummy variables with current group as reference
                cox_categorical_df = cox_categorical_base.copy()
                risk_dummies = pd.get_dummies(cox_categorical_df['PRSQuantile'], prefix='Risk', drop_first=False)
                risk_dummies = risk_dummies.drop(f'Risk_{ref_group}', axis=1)
                
                # Drop PRSQuantile and add dummies
                cox_categorical_df = cox_categorical_df.drop('PRSQuantile', axis=1)
                cox_categorical_df = pd.concat([cox_categorical_df, risk_dummies], axis=1)
                
                # Fit categorical Cox model
                cph_cat = CoxPHFitter()
                cph_cat.fit(cox_categorical_df, duration_col='DiseaseAfter', event_col='event_occurred')
                
                # Print summary for risk group comparisons only
                risk_cols = [col for col in cph_cat.summary.index if col.startswith('Risk_')]
                if risk_cols:
                    print(f"\nHazard Ratios vs {ref_group}:")
                    print(cph_cat.summary.loc[risk_cols, ['coef', 'exp(coef)', 'p', 'z']])
                
                # Store pairwise results
                if save_raw:
                    ci_cols = cph_cat.confidence_intervals_.columns  # e.g. ['95% lower-bound', '95% upper-bound']
                    for col in risk_cols:
                        comparison_group = col.replace('Risk_', '')
                        comparison_key = f"{comparison_group}_vs_{ref_group}"
                        raw_store['cox_ph']['categorical']['pairwise_comparisons'][comparison_key] = {
                            'hazard_ratio': cph_cat.hazard_ratios_[col],
                            'coef': cph_cat.summary.loc[col, 'coef'],
                            'exp_coef': cph_cat.summary.loc[col, 'exp(coef)'],
                            'p_value': cph_cat.summary.loc[col, 'p'],
                            'z_score': cph_cat.summary.loc[col, 'z'],
                            'confidence_interval_lower_exp_coef': np.exp(cph_cat.confidence_intervals_.loc[col, ci_cols[0]]),
                            'confidence_interval_upper_exp_coef': np.exp(cph_cat.confidence_intervals_.loc[col, ci_cols[1]])
                        }
                    
                    # Store concordance index (same for all models)
                    if raw_store['cox_ph']['categorical']['concordance_index'] is None:
                        raw_store['cox_ph']['categorical']['concordance_index'] = cph_cat.concordance_index_
                        print(f"\nConcordance Index: {cph_cat.concordance_index_:.4f}")
            
            # Create forest plot for hazard ratios (using Mid risk PRS as reference)
            print("\n--- Generating Forest Plot ---")
            
            # Re-fit with Mid risk PRS as reference for plotting
            cox_categorical_df_plot = cox_categorical_base.copy()
            risk_dummies_plot = pd.get_dummies(cox_categorical_df_plot['PRSQuantile'], prefix='Risk', drop_first=False)
            risk_dummies_plot = risk_dummies_plot.drop('Risk_Mid risk PRS', axis=1)
            cox_categorical_df_plot = cox_categorical_df_plot.drop('PRSQuantile', axis=1)
            cox_categorical_df_plot = pd.concat([cox_categorical_df_plot, risk_dummies_plot], axis=1)
            
            cph_plot = CoxPHFitter()
            cph_plot.fit(cox_categorical_df_plot, duration_col='DiseaseAfter', event_col='event_occurred')
            
            # Extract hazard ratios and CIs for risk groups only
            risk_cols_plot = [col for col in cph_plot.summary.index if col.startswith('Risk_')]
            if risk_cols_plot:
                hazard_ratios = []
                ci_lower = []
                ci_upper = []
                labels = []
                ci_cols_plot = cph_plot.confidence_intervals_.columns
                
                for col in risk_cols_plot:
                    hazard_ratios.append(cph_plot.summary.loc[col, 'exp(coef)'])
                    ci_lower.append(np.exp(cph_plot.confidence_intervals_.loc[col, ci_cols_plot[0]]))
                    ci_upper.append(np.exp(cph_plot.confidence_intervals_.loc[col, ci_cols_plot[1]]))
                    labels.append(col.replace('Risk_', ''))
                
                # Create forest plot
                fig, ax = plt.subplots(figsize=(8, max(4, len(labels)*0.8)), constrained_layout=True)
                
                y_positions = np.arange(len(labels))
                
                # Plot hazard ratios with error bars
                for i, (hr, ci_l, ci_u, label) in enumerate(zip(hazard_ratios, ci_lower, ci_upper, labels)):
                    ax.plot([hr], [i], 'o', color=res_colours[i % len(res_colours)], markersize=10, zorder=3)
                    ax.plot([ci_l, ci_u], [i, i], '-', color=res_colours[i % len(res_colours)], linewidth=2, zorder=2)
                
                # Add reference line at HR=1
                ax.axvline(x=1.0, color='gray', linestyle='--', linewidth=1.5, alpha=0.7, zorder=1)
                
                # Formatting
                ax.set_yticks(y_positions)
                ax.set_yticklabels(labels, fontsize=13)
                ax.set_xlabel('Hazard Ratio (95% CI)', fontsize=15)
                ax.set_title(f'Cox PH: Hazard Ratios vs Mid risk PRS\n{disease} [Sex: {sex}]', fontsize=16)
                ax.grid(axis='x', alpha=0.3, linestyle=':', zorder=0)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                
                # Add HR values as text annotations
                for i, (hr, ci_l, ci_u) in enumerate(zip(hazard_ratios, ci_lower, ci_upper)):
                    ax.text(ax.get_xlim()[1] * 0.98, i, f'{hr:.2f} ({ci_l:.2f}-{ci_u:.2f})', 
                           ha='right', va='center', fontsize=11, fontweight='bold')
                
                plt.xticks(fontsize=13)
                
                if args.save_plots:
                    covar_subdir = "covar_adjusted/" if args.adjust_prs_for_covars else ""
                    os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}{tag}cox_ph_forest"), exist_ok=True)
                    plt.savefig(f"{args.output_root}/plots/{covar_subdir}{tag}cox_ph_forest/bestfold_{disease.replace(' ', '-')}_{sex}.{save_format}", dpi=300, bbox_inches='tight')
                else:
                    plt.show()
                
                plt.close()
            
            # Test proportional hazards assumption with Mid risk as reference
            print("\n--- Proportional Hazards Assumption Test ---")
            cox_categorical_df = cox_categorical_base.copy()
            risk_dummies = pd.get_dummies(cox_categorical_df['PRSQuantile'], prefix='Risk', drop_first=False)
            risk_dummies = risk_dummies.drop('Risk_Mid risk PRS', axis=1)
            cox_categorical_df = cox_categorical_df.drop('PRSQuantile', axis=1)
            cox_categorical_df = pd.concat([cox_categorical_df, risk_dummies], axis=1)
            
            cph_cat_final = CoxPHFitter()
            cph_cat_final.fit(cox_categorical_df, duration_col='DiseaseAfter', event_col='event_occurred')
            
            try:
                ph_assumption_cat = cph_cat_final.check_assumptions(cox_categorical_df, p_value_threshold=0.05, show_plots=False)
                print("Proportional hazards assumption check completed.")
            except Exception as e:
                print(f"Warning: Could not test proportional hazards assumption: {e}")
            
            # Also store full summary from mid-reference model for compatibility
            if save_raw:
                raw_store['cox_ph']['categorical']['full_summary_mid_reference'] = cph_cat_final.summary.to_dict()
                raw_store['cox_ph']['categorical']['aic_partial'] = cph_cat_final.AIC_partial_
                raw_store['cox_ph']['categorical']['log_likelihood'] = cph_cat_final.log_likelihood_
            
            print(f"\n{'='*80}\n")

    if save_raw:
        return raw_store
    else:
        return None
    
# %% Other support functions

def getImprovedDisease(df, tag0, tag1):
        filtered_df = df[df['tag'].isin([tag0, tag1])]
        pivot_df = filtered_df.pivot(index='Disease', columns='tag', values='AUC_test').reset_index()
        result_df = pivot_df[pivot_df[tag0] < pivot_df[tag1]]
        return result_df['Disease'].tolist()

def getBestPRS(rds_pres_prefix, rds_pres_suffix, rds_tag_prs, tag_data, tag_prs, IIDs, res_method=None, resDF=None):
    if resDF is not None:
        best_prs_latent = resDF.iloc[0].Latent
    else:
        best_prs_latent = res_method.iloc[res_method['AUC_test'].argmax()].name
    pth_prs = f"{rds_pres_prefix}{best_prs_latent.split('PRS:')[-1]}{rds_pres_suffix}"

    data = pyreadr.read_r(pth_prs.replace(rds_tag_prs, tag_data))[None]
    prs = data[['IID']]
    r_obj = read_rds(pth_prs)
    prs["predicted"] = r_obj['data'][r_obj['attributes']['names']['data'].index(tag_prs)]['data']
    prs.set_index("IID", inplace=True)
    prs = prs[prs.index.isin(IIDs)]

    return prs

def select_values(row, base_cols, suffixes):
    for suffix in suffixes:
        if all(not pd.isna(row[f"{col}.{suffix}.0"]) for col in base_cols):
            return [row[f"{col}.{suffix}.0"] for col in base_cols] + [suffix]
    return [pd.NA] * len(base_cols)

def get_disease_info(disease_path, raw_baseline_path, info_assessment_centre_path, cutoff_date):
    diseases = pd.read_csv(disease_path, low_memory=False).set_index("eid")

    raw_baseline = pd.read_table(raw_baseline_path).set_index("f.eid")
    raw_baseline = raw_baseline[[c for c in raw_baseline.columns if "f.20116." in c or "f.21001" in c]]
    assessment_centre = pd.read_table(info_assessment_centre_path).set_index("f.eid")
    assessment_centre = assessment_centre[[c for c in assessment_centre.columns if "f.53." in c]]
    df_unprocessed = diseases.merge(raw_baseline, left_index=True, right_index=True).merge(assessment_centre, left_index=True, right_index=True) 
    df_unprocessed.index.names = ['eid']
    df_unprocessed[['f.20116', 'f.21001', 'f.53', 'selected_instance']] = df_unprocessed.apply(lambda row: pd.Series(select_values(row, base_cols=['f.20116', 'f.21001', 'f.53'], suffixes=['0', '1', '2', '3'])), axis=1)
    df_unprocessed = df_unprocessed.drop([c for c in df_unprocessed.columns if "f.20116." in c or "f.21001." in c or "f.53." in c], axis=1) #we will remove now the instance-wise columns
    df_unprocessed['date'] = pd.to_datetime(df_unprocessed['date'], format='%Y-%m-%d', errors='coerce')
    df_unprocessed['f.53'] = pd.to_datetime(df_unprocessed['f.53'], format='%Y-%m-%d', errors='coerce')

    #remove subjects with "future" diseases
    df_unprocessed = df_unprocessed[((df_unprocessed['f.53'].isna()) | (df_unprocessed['f.53'] <= pd.to_datetime(cutoff_date))) & ((df_unprocessed['date'].isna()) | (df_unprocessed['date'] <= pd.to_datetime(cutoff_date)))]

    df_unprocessed['DiseaseAfter'] = (df_unprocessed['date'] - df_unprocessed['f.53']).dt.days
    df_unprocessed.rename(columns={'f.20116': 'CAT_Smoking', 'f.21001': 'BMI', 'f.53': 'BaselineDate'}, inplace=True)

    df_unprocessed = df_unprocessed[df_unprocessed['CAT_Smoking'] != -3].copy()

    df_unprocessed = df_unprocessed.sort_values(by='date', na_position='last')
    df_unprocessed = df_unprocessed.reset_index()
    return df_unprocessed[~df_unprocessed.duplicated(subset=['summary', 'eid'], keep='first')].sort_values(by='eid').set_index('eid')

# %% Support functions to make the script modular (and to make it compatible with sex stratified analyses)

def load_and_split_data(df, covar=None):
    if bool(covar):
        df = df.join(covar, how='inner')
    male_data = df[df['Sex'] == 1]
    female_data = df[df['Sex'] == 0]
    return df, male_data, female_data

def logOR(data_test, col_y, pred_probs_test):
    try:
        data_test['predicted_prob'] = pred_probs_test
        data_test['quantile'] = pd.qcut(data_test['predicted_prob'], 5, labels=False, duplicates='drop') + 1
        odds_q3 = (data_test[data_test['quantile'] == 3][col_y].sum()) / \
                (data_test[data_test['quantile'] == 3][col_y].count() - data_test[data_test['quantile'] == 3][col_y].sum())
        odds_q5 = (data_test[data_test['quantile'] == 5][col_y].sum()) / \
                (data_test[data_test['quantile'] == 5][col_y].count() - data_test[data_test['quantile'] == 5][col_y].sum())
        return np.log(odds_q5 / odds_q3)
    except:
        return np.nan

def getF1(label, pred_prob, method="at50"):
    match method:
        case "otsu":        
            threshold = threshold_otsu(pred_prob.values)
        case "at50":
            threshold = 0.5
        case _:
            sys.exit(f"Method {method} for getF1 not implemented")
    pred = pred_prob > threshold
    return f1_score(label, pred)
    


@plots_and_results_by_sex
def get_scores(resDF, fold, method, res_type, diseaseDF, IIDs, sex, dis_col="BinCAT_Disease"):
    df = diseaseDF.join(resDF, how='inner')
    pred_col = resDF.name if type(resDF) is pd.Series else resDF.columns[0]
    df = df[df.index.isin(IIDs)]
    auc = roc_auc_score(df[dis_col], df[pred_col])
    logOR_value = logOR(df, pred_col, df[pred_col])
    f1_test = getF1(df[dis_col], df[pred_col], method="at50")
    return {"AUC_test": auc, "logOR_test": logOR_value, "F1_test": f1_test, "Sex": sex, "fold": fold, "method": method, "res_type": res_type}

@plots_and_results_by_sex
def get_scores_pancohort(resDF, fold, method, res_type, IIDs, sex):
    if type(resDF) is pd.Series: 
        pred_col = resDF.name
        df = pd.DataFrame(resDF[resDF.index.isin(IIDs)].rename(pred_col)) #the rename call is required for older pandas versions (e.g. 1.4.4)
    else:
        df = resDF[resDF.index.isin(IIDs)]
        pred_col = resDF.columns[0] 
    logOR_value = logOR(df, pred_col, df[pred_col])
    return {"logOR_test_pancohort": logOR_value, "Sex": sex, "fold": fold, "method": method, "res_type": res_type}

# %%
if __name__ == "__main__":
    parser = getARGSParser()
    args, _ = parser.parse_known_args()

    args.colour_box = args.colour_box.split(",") if bool(args.colour_box) else colours
    args.colour_prevalence = args.colour_prevalence.split(",") if bool(args.colour_prevalence) else colours
    args.colour_displots = args.colour_displots.split(",") if bool(args.colour_displots) else colours
    
    # Parse quick run arguments
    if args.quick_run:
        if args.quick_run_sex:
            args.quick_run_sex = [s.strip() for s in args.quick_run_sex.split(",")]
            # Validate sex groups
            valid_sex = ['Both', 'Female', 'Male']
            for s in args.quick_run_sex:
                if s not in valid_sex:
                    raise ValueError(f"Invalid sex group '{s}'. Must be one of: {valid_sex}")
        else:
            args.quick_run_sex = ['Both', 'Female', 'Male']  # Default: all three
        
        if args.quick_run_diseases:
            args.quick_run_diseases = [d.strip() for d in args.quick_run_diseases.split(",")]
        else:
            args.quick_run_diseases = []  # Empty means all diseases
    else:
        args.quick_run_sex = ['Both', 'Female', 'Male']  # Default when not in quick run mode
        args.quick_run_diseases = []

    print(args)

    covars = pd.read_table(args.ext_covar, low_memory=False)
    
    # Parse covariate columns
    if args.covar_cont_cols:
        args.covar_cont_cols = [c.strip() for c in args.covar_cont_cols.split(",")]
    else:
        # Default: use Age and all PC columns
        args.covar_cont_cols = ['Age'] + [c for c in covars.columns if c.startswith('PC')]
    
    if args.covar_cat_cols:
        args.covar_cat_cols = [c.strip() for c in args.covar_cat_cols.split(",")]
    else:
        args.covar_cat_cols = []
    
    # Extract male/female IIDs using Sex column
    if 'Sex' in covars.columns:
        IIDs_male = covars[covars.Sex==1].IID.to_list()
        IIDs_female = covars[covars.Sex==0].IID.to_list()
    else:
        # If Sex not available, use all IIDs for both
        IIDs_male = covars.IID.to_list()
        IIDs_female = covars.IID.to_list()

    # %% define the interesting comparisons [TODO: make it configurable. Currently, it's a "comment-out"-based approach]
    comparisons = [
        ('GLM', 'covar', 'Baseline', 1),
        ('Lasso', 'covar', 'Lasso Baseline', 0),

        ('GLM', 'covarNorm', 'Baseline (Normalised)', 0),
        ('Lasso', 'covarNorm', 'Lasso Baseline (Normalised)', 0),

        # ('GLM', 'nonPCCovar', 'non-PC Baseline', 0),
        # ('Lasso', 'nonPCCovar', 'Lasso non-PC Baseline', 0),

        ('GLM', 'singlePRS', 'max(Single PRS)', 0),

        ('GLM', 'singlePRSCovar', 'max(Single PRS + Baseline)', 1),
        ('Lasso', 'singlePRSCovar', 'max(Lasso(Single PRS + Baseline))', 0),
        
        ('GLM', 'singlePRSCovarNorm', 'max(norm(Single PRS + Baseline))', 0),
        ('Lasso', 'singlePRSCovarNorm', 'max(Lasso(norm(Single PRS + Baseline)))', 0),

        ('GLM', 'multiPRS', 'Multi PRS (GLM)', 0),
        ('Lasso', 'multiPRS', 'Multi PRS', 0),
        ('GLM', 'multiPRSCovar', 'Multi PRS + Baseline (GLM)', 0),
        ('Lasso', 'multiPRSCovar', 'Multi PRS + Baseline', 1),
        # ('GLM', 'multiPRSnonPCCovar', 'Multi PRS + non-PC Baseline (GLM)', 0),
        # ('Lasso', 'multiPRSnonPCCovar', 'Multi PRS + non-PC Baseline', 0),

        ('GLM', 'multiPRSNorm', 'Normalised Multi PRS (GLM)', 0),
        ('Lasso', 'multiPRSNorm', 'Normalised Multi PRS', 0),
        ('GLM', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (GLM)', 0),
        ('Lasso', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline', 0),
        # ('GLM', 'multiPRSNormnonPCCovar', 'Normalised Multi PRS + non-PC Baseline (GLM)', 0),
        # ('Lasso', 'multiPRSNormnonPCCovar', 'Normalised Multi PRS + non-PC Baseline', 0),

        ('LassoSteps0', 'multiPRSCovar', 'Multi PRS + Baseline (LassoSteps0)', 0),
        ('LassoSteps0', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (LassoSteps0)', 0),

        ('LassoSteps1', 'multiPRSCovar', 'Multi PRS + Baseline (LassoSteps1)', 0),
        ('LassoSteps1', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (LassoSteps1)', 0),

        ('LassoSteps2', 'multiPRSCovar', 'Multi PRS + Baseline (LassoSteps2)', 0),
        ('LassoSteps2', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (LassoSteps2)', 0),

        ('LassoSteps3', 'multiPRSCovar', 'Multi PRS + Baseline (LassoSteps3)', 0),
        ('LassoSteps3', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (LassoSteps3)', 0),

        # ('ElasticNet', 'multiPRS', 'Multi PRS', 0),
        # ('ElasticNet', 'multiPRSCovar', 'Multi PRS + Baseline', 0),
        # # ('ElasticNet', 'multiPRSnonPCCovar', 'Multi PRS + non-PC Baseline', 0),

        # ('ElasticNet', 'multiPRSNorm', 'Normalised Multi PRS', 0),
        # ('ElasticNet', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline', 0),
        # # ('ElasticNet', 'multiPRSNormnonPCCovar', 'Normalised Multi PRS + non-PC Baseline', 0),

        # ('XGBoost', 'multiPRS', 'Multi PRS', 0),
        # ('XGBoost', 'multiPRSCovar', 'Multi PRS + Baseline', 0),
        # ('ElasticNet', 'multiPRSnonPCCovar', 'Multi PRS + non-PC Baseline', 0),

        # ('XGBoost', 'multiPRSNorm', 'Normalised Multi PRS', 0),
        # ('XGBoost', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline', 0),
        # ('ElasticNet', 'multiPRSNormnonPCCovar', 'Normalised Multi PRS + non-PC Baseline', 0),
    ]

    all_combinations = list(itertools.combinations(comparisons, 2))
    interesting_combinations = {(method, res_type) for method, res_type, _, _ in comparisons}
    tag_lookup = {(method, res_type): tag for method, res_type, tag, _ in comparisons}
    tag_order = [tag for _, _, tag, _ in comparisons]

    # %% obtain the results' pickles
    results = glob(f"{args.output_root}/*_raw_results.pkl")    


    # %% create summary of the scores and perform statistical tests

    if args.obtain_summary:
        collect_df = []
        collect_summary_df = []
        collect_stats = []

        with open(f"{args.output_root}/sumscores.txt", 'w') as f:
            for r in tqdm(results):
                d = os.path.basename(r).replace('_raw_results.pkl', '')
                f.write(f"\n\n{d}:-------------------------------\n")

                disease = pd.read_csv(f'{args.disease_root}/{d}.csv', low_memory=False, index_col="IID")
                disease["BinCAT_Disease"] = disease["BinCAT_Disease"].astype(int)

                with open(r, "rb") as pkl:
                    res = pickle.load(pkl)

                scores = []
                scores_pancohort = []
                for fold in res.keys():
                    for method in res[fold].keys():
                        if method == 'IDs':
                            continue
                        for res_type in res[fold][method].keys():
                            best_auc = {"Both": 0.0, "Female": 0.0, "Male": 0.0}
                            best_scores = {}
                            if "single" in res_type.lower():
                                for latent in res[fold][method][res_type]['pred_probs_test'].keys():
                                    scr = get_scores(resDF=res[fold][method][res_type]['pred_probs_test'][latent], diseaseDF=disease, IIDs_female=IIDs_female, IIDs_male=IIDs_male, fold=fold, method=method, res_type=res_type)
                                    for s in scr:
                                        if s['AUC_test'] > best_auc[s["Sex"]]:
                                            best_auc[s["Sex"]] = s['AUC_test']
                                            best_scores[s["Sex"]] = s
                                            best_scores[s["Sex"]]["Latent"] = latent
                                scores.append(list(best_scores.values()))
                                scores_pancohort.append(get_scores_pancohort(resDF=res[fold][method][res_type]['pred_probs_test_pancohort'][best_scores["Both"]["Latent"]], IIDs_female=IIDs_female, IIDs_male=IIDs_male, fold=fold, method=method, res_type=res_type))
                            else:
                                scores.append(get_scores(resDF=res[fold][method][res_type]['pred_probs_test'], diseaseDF=disease, IIDs_female=IIDs_female, IIDs_male=IIDs_male, fold=fold, method=method, res_type=res_type))
                                scores_pancohort.append(get_scores_pancohort(resDF=res[fold][method][res_type]['pred_probs_test_pancohort'], IIDs_female=IIDs_female, IIDs_male=IIDs_male, fold=fold, method=method, res_type=res_type))     
                scores = pd.DataFrame([score for sublist in scores for score in sublist])
                scores_pancohort = pd.DataFrame([score for sublist in scores_pancohort for score in sublist])
                df = pd.merge(scores, scores_pancohort, on=["Sex", "fold", "method", "res_type"])

                df['Disease'] = d.replace("-", " ")
                df['nPatients'] = len(disease) // 2

                df = df[df.apply(lambda row: (row['method'], row['res_type']) in interesting_combinations, axis=1)]
                df['tag'] = df.apply(lambda row: tag_lookup.get((row['method'], row['res_type'])), axis=1)
                df['tag'] = pd.Categorical(df['tag'], categories=tag_order, ordered=True)
                collect_df.append(df)

                summary_df = getSummaryDF(df, pancohort=args.is_pancohort)
                f.write("\nSummary scores:-")
                f.write(summary_df.to_string(index=False))
                summary_df['Disease'] = d.replace("-", " ")
                summary_df['nPatients'] = len(disease) // 2
                collect_summary_df.append(summary_df)

                summary_stats = compareSummaryPairs(df=df, all_combinations=all_combinations, disease=d.replace("-", " "), pancohort=args.is_pancohort)
                summary_stats = pd.concat(summary_stats)
                
                auc_delong = compareAUCPairs(res=res, all_combinations=all_combinations, disease=d.replace("-", " "), diseaseDF=disease, IIDs_female=IIDs_female, IIDs_male=IIDs_female, resDF=df)
                auc_delong = pd.concat(auc_delong)

                stats_combined = pd.merge(auc_delong, summary_stats, on=["Disease", "Comparison", "Sex", "Method0", "Method1"]) #Method0 and Method1, along with Comparison, are redundant. But, I'm keeping them just to avoid those columns appearing twice with _x and _y suffixes. 
                stats_combined['nPatients'] = len(disease) // 2
                collect_stats.append(stats_combined)

                stats2print = stats_combined[['Comparison', 'Sex', 'p_value_DeLong_AUC']].copy()
                stats2print['is Sig'] = stats2print['p_value_DeLong_AUC'].apply(lambda x: 'Sig' if x < 0.05 else 'Not Sig')
                f.write("\n\nDeLong's p-values:-")
                f.write(stats2print.to_string(index=False))

                f.write("\n\n-----------------------------------\n\n")

        collect_df = pd.concat(collect_df)
        collect_summary_df = pd.concat(collect_summary_df)
        collect_stats = pd.concat(collect_stats)

        collect_df.to_csv(f"{args.output_root}/combined_results.tsv", sep="\t", index=False)
        collect_summary_df.to_csv(f"{args.output_root}/combined_summary_results.tsv", sep="\t", index=False)
        collect_stats.to_csv(f"{args.output_root}/combined_stats_sig.tsv", sep="\t", index=False)

    else:
        collect_df = pd.read_table(f"{args.output_root}/combined_results.tsv", low_memory=False)
        collect_summary_df = pd.read_table(f"{args.output_root}/combined_summary_results.tsv", low_memory=False)
        collect_stats = pd.read_table(f"{args.output_root}/combined_stats_sig.tsv", low_memory=False)

    # %%
    if bool(args.drop_dis):
        print(f"Dropping {args.drop_dis} from the analysis")
        collect_df = collect_df[~collect_df.Disease.isin(args.drop_dis.split(","))]
        collect_summary_df = collect_summary_df[~collect_summary_df.Disease.isin(args.drop_dis.split(","))]
        collect_stats = collect_stats[~collect_stats.Disease.isin(args.drop_dis.split(","))]

    if args.dis_plots_rawPRS:
        collect_df_rawPRS = collect_df[collect_df.tag == "max(Single PRS)"]

    if args.drop_comparisons:
        selected_comparisons = []
        for c in comparisons:
            if c[3] == 1:
                selected_comparisons.append(c)
        interesting_combinations = {(method, res_type) for method, res_type, _, _ in selected_comparisons}
        tag_lookup = {(method, res_type): tag for method, res_type, tag, _ in selected_comparisons}
        tag_order = [tag for _, _, tag, _ in selected_comparisons]
        collect_df = collect_df[collect_df.tag.isin(tag_order)]
        collect_summary_df = collect_summary_df[collect_summary_df.tag.isin(tag_order)]
        collect_stats = collect_stats[collect_stats.Method0.isin(tag_order) & collect_stats.Method1.isin(tag_order)]  
        comparisons = selected_comparisons

    # %% print the pairwise improvements

    if args.obtain_pairwise_improvements:    
        combined_df = combineDFs(collect_stats, collect_summary_df)
        tab_summary = summary_table(combined_df)
        for metric in ["AUC", "logOR", "logOR_Pancohort", "F1"]:
            create_excel_summary(tab_summary, metric, f"{args.output_root}/summary_{metric}.xlsx")
        combined_df.to_csv(f"{args.output_root}/summary_combined.tsv", sep="\t", index=False)
        tab_summary.to_csv(f"{args.output_root}/summary_table.tsv", sep="\t", index=False)

    # %% "improve" the dataframes for plotting
    collect_df.Disease = collect_df.Disease.apply(lambda x: transform_text(x, "sentence_case"))
    if args.dis_plots_rawPRS:
        collect_df_rawPRS.Disease = collect_df_rawPRS.Disease.apply(lambda x: transform_text(x, "sentence_case"))

    # %% sort the diseases based on the AUC
    sorted_diseases = collect_df[(collect_df['res_type'] == 'singlePRSCovar') & (collect_df.Sex == 'Both')].groupby('Disease')['AUC_test'].mean().sort_values(ascending=False).index
    match (args.sort_mode_top_dis):
        case "0":
            pass
        case "1":
            sorted_diseases = sorted(sorted_diseases)
        case _ :
            print("Using the provided custom order...")
            sorted_diseases = args.sort_mode_top_dis.split(",")

    if args.n_top_dis > 0:
        sorted_diseases = sorted_diseases[:args.n_top_dis]

    if args.split_top_dis_half:
        top_diseases = sorted_diseases[:(len(sorted_diseases)//2)]
        bottom_diseases = sorted_diseases[(len(sorted_diseases)//2):]
    else:
        top_diseases = sorted_diseases
        bottom_diseases = []

    df_top = collect_df[collect_df['Disease'].isin(top_diseases)].copy()
    df_top['Disease'] = pd.Categorical(df_top['Disease'], categories=top_diseases, ordered=True)

    if bool(bottom_diseases):
        df_bottom = collect_df[collect_df['Disease'].isin(bottom_diseases)].copy()
        df_bottom['Disease'] = pd.Categorical(df_bottom['Disease'], categories=bottom_diseases, ordered=True)

    if args.save_raw:
        raw_store_base = "covar_adjusted/raw_store" if args.adjust_prs_for_covars else "raw_store"
        os.makedirs(os.path.join(args.output_root, "plots", raw_store_base), exist_ok=True)

    # %% Create box plots for AUC_test, top N//2 diseases and bottom N//2 diseases seperately

    if args.plot_box_AUC:
        os.makedirs(os.path.join(args.output_root, "plots", "auc"), exist_ok=True)

        plot_box(df_top, x='Disease', y='AUC_test', y_tag='AUC', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/auc/top_{len(top_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_AUC)
        if bool(bottom_diseases):
            plot_box(df_bottom, x='Disease', y='AUC_test', y_tag='AUC', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/auc/bottom_{len(bottom_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_AUC)
    
    # %% Create box plots for F1_test, top N//2 diseases and bottom N//2 diseases seperately

    if args.plot_box_F1:
        os.makedirs(os.path.join(args.output_root, "plots", "f1"), exist_ok=True)

        plot_box(df_top, x='Disease', y='F1_test', y_tag='F1', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/f1/top_{len(top_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_F1)
        if bool(bottom_diseases):
            plot_box(df_bottom, x='Disease', y='F1_test', y_tag='F1', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/f1/bottom_{len(bottom_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_F1)
    
    # %% Create box plots for logOR_test, top N//2 diseases and bottom N//2 diseases seperately

    if args.plot_box_logOR:
        os.makedirs(os.path.join(args.output_root, "plots", "logOR"), exist_ok=True)

        plot_box(df_top, x='Disease', y='logOR_test', y_tag='logOR', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/logOR/top_{len(top_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_logOR)
        if bool(bottom_diseases):
            plot_box(df_bottom, x='Disease', y='logOR_test', y_tag='logOR', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/logOR/bottom_{len(bottom_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_logOR)
    
    # %% Create box plots for logOR_test_pancohort, top N//2 diseases and bottom N//2 diseases seperately
    if args.is_pancohort and args.plot_box_logOR:
        os.makedirs(os.path.join(args.output_root, "plots", "logOR_pancohort"), exist_ok=True)

        plot_box(df_top, x='Disease', y='logOR_test_pancohort', y_tag='logOR (Pancohort)', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/logOR_pancohort/top_{len(top_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_logOR_pancohort)
        if bool(bottom_diseases):
            plot_box(df_bottom, x='Disease', y='logOR_test_pancohort', y_tag='logOR (Pancohort)', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/logOR_pancohort/bottom_{len(bottom_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_logOR_pancohort)

    # %% Create prevalence vs PRS scores plots for the pancohort, for the best single latent PRS

    if args.plot_prevalence_prs_pancohort:

        if args.plot_prevalence_prs_pancohort:
            prevalence_prs_subdir = "covar_adjusted/prevalence_prs_pancohort" if args.adjust_prs_for_covars else "prevalence_prs_pancohort"
            os.makedirs(os.path.join(args.output_root, "plots", prevalence_prs_subdir), exist_ok=True)

        raw_res_store = []
        for dis in sorted_diseases:
            r = f"{args.output_root}/{dis.replace(' ', '-').lower()}_raw_results.pkl"

            with open(r, "rb") as pkl:
                res = pickle.load(pkl)
        
            best_AUC = 0
            best_prs_latent = None
            for fold in res.keys():
                _, res_method = getIndividualProbs(res[fold]['GLM']['singlePRS'], 'singlePRS', IIDs=IIDs_female+IIDs_male) #Maybe we can think about also doing it in a sex sex-stratified manner?
                if res_method['AUC_test'] > best_AUC:
                    best_AUC = res_method['AUC_test']
                    best_prs_latent = res_method.name.split("PRS:")[-1]
                
            pth_prs = f"{args.rds_pres_prefix}{best_prs_latent}{args.rds_pres_suffix}"

            data = pyreadr.read_r(pth_prs.replace(args.rds_tag_prs, args.tag_data))[None]
            prs = data[['IID']]
            r_obj = read_rds(pth_prs)
            prs["PRS"] = r_obj['data'][r_obj['attributes']['names']['data'].index(args.tag_prs)]['data']
            prs.set_index("IID", inplace=True)

            # Adjust PRS for covariates using linear regression (regress out covariate effects)
            if args.adjust_prs_for_covars and (args.covar_cont_cols or args.covar_cat_cols):
                merge_cols = ['IID'] + (args.covar_cont_cols if args.covar_cont_cols else []) + (args.covar_cat_cols if args.covar_cat_cols else [])
                merge_cols = [c for c in merge_cols if c in covars.columns]
                prs_with_covars = prs.merge(covars[merge_cols], left_index=True, right_on='IID', how='inner').set_index('IID')
                available_cont_covars = [c for c in args.covar_cont_cols if c in prs_with_covars.columns] if args.covar_cont_cols else []
                available_cat_covars = [c for c in args.covar_cat_cols if c in prs_with_covars.columns] if args.covar_cat_cols else []
                
                if available_cont_covars or available_cat_covars:
                    X_parts = []
                    if available_cont_covars:
                        X_cont = prs_with_covars[available_cont_covars].fillna(prs_with_covars[available_cont_covars].median())
                        X_parts.append(X_cont)
                    if available_cat_covars:
                        X_cat = pd.get_dummies(prs_with_covars[available_cat_covars], drop_first=True)
                        X_parts.append(X_cat)
                    X_covars = pd.concat(X_parts, axis=1)
                    lr = LinearRegression()
                    lr.fit(X_covars, prs_with_covars['PRS'])
                    prs_with_covars['PRS_adjusted'] = prs_with_covars['PRS'] - lr.predict(X_covars)
                    prs = prs_with_covars[['PRS_adjusted']].rename(columns={'PRS_adjusted': 'PRS'})

            d = os.path.basename(r).replace('_raw_results.pkl', '')
            disease = pd.read_csv(f'{args.disease_root}/{d}.csv', low_memory=False, index_col="IID")
            combined = prs.join(disease[['BinCAT_Disease']])
            combined.fillna(0, inplace=True)
            combined["BinCAT_Disease"] = combined["BinCAT_Disease"].astype(int)

            if args.save_plots:
                pth = f"{args.output_root}/plots/{prevalence_prs_subdir}/{d.replace(' ', '-')}.{args.save_format}"
            else:
                pth = ""

            d = d.replace("-", " ")
            if args.sex_stratified:
                data.IID = data.IID.astype(combined.index.dtype)
                IIDs_all = list(combined.index)
                IIDs_female = list(set(IIDs_all).intersection(set(list(data[data.Sex == "0"].IID))))
                IIDs_male = list(set(IIDs_all).intersection(set(list(data[data.Sex == "1"].IID))))
                combined.loc[IIDs_female, 'PRS_Female'] = combined.loc[IIDs_female, 'PRS']
                combined.loc[IIDs_male, 'PRS_Male'] = combined.loc[IIDs_male, 'PRS']
                pth = pth.replace(f".{args.save_format}", f"_sex_stratified.{args.save_format}")
                _raw_entry = plot_prevalence(combined, res_keys=['PRS_Female', 'PRS', 'PRS_Male'], res_labels=['Female', 'Both', 'Male'], res_colours=args.colour_prevalence[:3], res_fmts=markers[:3], col_disease='BinCAT_Disease', plot_title=f'max(Single PRS) Risk scores: {d}', individual_prev_line=True, save_path=pth, save_raw=args.save_raw)
            else:
                _raw_entry = plot_prevalence(combined, res_keys=['PRS'], res_labels=['max(Single PRS)'], res_colours=['blue'], res_fmts=['o'], col_disease='BinCAT_Disease', plot_title=f'Risk scores: {d}', save_path=pth, save_raw=args.save_raw)
            # Annotate the raw_store entry with the disease so downstream TSV exports are self-describing
            if isinstance(_raw_entry, dict):
                _raw_entry['Disease'] = d
                _raw_entry['sex_stratified'] = bool(args.sex_stratified)
            raw_res_store.append(_raw_entry)

        if args.save_raw:
            raw_store_subdir = "covar_adjusted/raw_store" if args.adjust_prs_for_covars else "raw_store"
            os.makedirs(os.path.join(args.output_root, "plots", raw_store_subdir), exist_ok=True)
            with open(f'{os.path.join(args.output_root, "plots", raw_store_subdir)}/raw_prevalence_prs_pancohort.pkl', "wb") as pkl:
                pickle.dump(raw_res_store, pkl)

    # %% Create prevalence vs probability plots for the pancohort, for the best fold
    
    raw_res_store = []
    if args.is_pancohort and args.plot_prevalence_prob_pancohort:
        prevalence_prob_subdir = "covar_adjusted/prevalence_prob_pancohort" if args.adjust_prs_for_covars else "prevalence_prob_pancohort"
        os.makedirs(os.path.join(args.output_root, "plots", prevalence_prob_subdir), exist_ok=True)

        for dis in sorted_diseases:
            r = f"{args.output_root}/{dis.replace(' ', '-').lower()}_raw_results.pkl"

            with open(r, "rb") as pkl:
                res = pickle.load(pkl)

            d = os.path.basename(r).replace('_raw_results.pkl', '')
            diseaseDF = pd.read_csv(f'{args.disease_root}/{d}.csv', low_memory=False, index_col="IID")

            raw_res_store.append(plot_prevalence_prob_pancohort(res, dis, diseaseDF, collect_df, IIDs_female=IIDs_female, IIDs_male=IIDs_male, comparisons=comparisons, save_format=args.save_format))

        if args.save_raw:
            raw_store_subdir = "covar_adjusted/raw_store" if args.adjust_prs_for_covars else "raw_store"
            os.makedirs(os.path.join(args.output_root, "plots", raw_store_subdir), exist_ok=True)
            with open(f'{os.path.join(args.output_root, "plots", raw_store_subdir)}/raw_prevalence_prob_pancohort.pkl', "wb") as pkl:
                pickle.dump(raw_res_store, pkl)
        
    # %% Create prevalence vs probability plots for the pancohort, for the best fold

    if args.plot_cum_disease_burden or args.plot_cum_hazard or args.plot_KM_survival or args.plot_cox_ph:

        covar_subdir = "covar_adjusted/" if args.adjust_prs_for_covars else ""

        if args.plot_cum_disease_burden:
            os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}cumulative_disease_burden"), exist_ok=True)
            if args.dis_plots_rawPRS:
                os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}rawPRS_cumulative_disease_burden"), exist_ok=True)
        if args.plot_cum_hazard:
            os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}cumulative_hazard"), exist_ok=True)
            if args.dis_plots_rawPRS:
                os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}rawPRS_cumulative_hazard"), exist_ok=True)
        if args.plot_KM_survival:
            os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}KM_survival"), exist_ok=True)
            if args.dis_plots_rawPRS:
                os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}rawPRS_KM_survival"), exist_ok=True)
        if args.plot_cox_ph:
            os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}cox_ph_forest"), exist_ok=True)
            if args.dis_plots_rawPRS:
                os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}rawPRS_cox_ph_forest"), exist_ok=True)

        if args.reprocess_raw_diseases or not os.path.isfile(f"{args.output_root}/raw_diseases.tsv"):
            raw_diseases = get_disease_info(disease_path=args.raw_disease_path, raw_baseline_path=args.raw_baseline_path, info_assessment_centre_path=args.raw_centre_info_path, cutoff_date=args.dis_plots_cutoff_date)
            raw_diseases = raw_diseases.join(covars.set_index("IID")['Sex'], how='inner')
            raw_diseases = add_grouped_diseases(raw_diseases, cardiac_mappings, "healthy")
            raw_diseases.to_csv(f"{args.output_root}/raw_diseases.tsv", sep="\t")
        else:
            print("Existing raw_diseases.tsv found! Loading...")
            raw_diseases = pd.read_table(f"{args.output_root}/raw_diseases.tsv", low_memory=False).set_index("eid")
            raw_diseases = add_grouped_diseases(raw_diseases, cardiac_mappings, "healthy") #TODO: remove it
        
        if args.save_raw:
            os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}raw_displots"), exist_ok=True)
            os.makedirs(os.path.join(args.output_root, "plots", f"{covar_subdir}raw_store"), exist_ok=True)
            raw_res_store = []
        
        # Filter diseases for quick run mode
        if args.quick_run and args.quick_run_diseases:
            # Convert sorted_diseases to list if it's an Index
            diseases_list = list(sorted_diseases)
            # Filter to only requested diseases (case-sensitive exact match)
            filtered_diseases = [d for d in diseases_list if d in args.quick_run_diseases]
            # Warn about any requested diseases not found
            not_found = set(args.quick_run_diseases) - set(filtered_diseases)
            if not_found:
                print(f"\nWARNING: The following diseases were requested but not found: {', '.join(not_found)}")
            if filtered_diseases:
                print(f"\nQuick Run Mode: Processing {len(filtered_diseases)}/{len(diseases_list)} diseases: {', '.join(filtered_diseases)}")
                sorted_diseases = filtered_diseases
            else:
                print("\nWARNING: No requested diseases found in sorted_diseases. Skipping disease plots.")
                sorted_diseases = []

        for dis in sorted_diseases:
            r = f"{args.output_root}/{dis.replace(' ', '-').lower()}_raw_results.pkl"

            with open(r, "rb") as pkl:
                res = pickle.load(pkl)
            
            if bool(args.dis_plots_mod):
                # Pass sex_groups if in quick run mode
                plot_kwargs = {
                    'res': res, 
                    'disease': dis, 
                    'diseaseDF': raw_diseases, 
                    'collect_df': collect_df, 
                    'args': args, 
                    'covars': covars, 
                    'IIDs_female': IIDs_female, 
                    'IIDs_male': IIDs_male, 
                    'model_type': args.dis_plots_mod.split(',')[0], 
                    'model': args.dis_plots_mod.split(',')[1], 
                    'PRS_col': "predicted", 
                    'ending_year': args.dis_plots_upto_Nyear, 
                    'use_raw_PRS': False, 
                    'res_colours': args.colour_displots, 
                    'save_format': args.save_format, 
                    'save_raw': args.save_raw
                }
                if args.quick_run:
                    plot_kwargs['sex_groups'] = args.quick_run_sex
                
                result = plot_dis_plots(**plot_kwargs)
                if result is not None:
                    raw_res_store.append(result)
            
                if args.dis_plots_rawPRS:
                    # Pass sex_groups if in quick run mode
                    plot_kwargs_raw = {
                        'res': res, 
                        'disease': dis, 
                        'diseaseDF': raw_diseases, 
                        'collect_df': collect_df_rawPRS, 
                        'args': args, 
                        'covars': covars, 
                        'IIDs_female': IIDs_female, 
                        'IIDs_male': IIDs_male, 
                        'model_type': "GLM", 
                        'model': "singlePRS", 
                        'PRS_col': "predicted", 
                        'ending_year': args.dis_plots_upto_Nyear, 
                        'use_raw_PRS': True, 
                        'res_colours': args.colour_displots, 
                        'save_format': args.save_format, 
                        'save_raw': args.save_raw
                    }
                    if args.quick_run:
                        plot_kwargs_raw['sex_groups'] = args.quick_run_sex
                    
                    result_raw = plot_dis_plots(**plot_kwargs_raw)
                    if result_raw is not None:
                        raw_res_store.append(result_raw)

        if args.save_raw:
            # Build filename suffix based on active plots (only in quick run mode)
            filename = "raw_displots"
            if args.quick_run:
                plot_tags = []
                if args.plot_cum_disease_burden:
                    plot_tags.append("cumBurden")
                if args.plot_cum_hazard:
                    plot_tags.append("cumHaz")
                if args.plot_KM_survival:
                    plot_tags.append("KM")
                if args.plot_cox_ph:
                    plot_tags.append("coxPH")
                
                # Only add suffix if it's a subset of plots (not all 4)
                if plot_tags and len(plot_tags) < 4:
                    filename = f"raw_displots_{'_'.join(plot_tags)}"
                    print(f"\nQuick Run Mode: Saving raw_store to {filename}.pkl (tagged to avoid overwrites)")
            
            with open(f'{os.path.join(args.output_root, "plots", f"{covar_subdir}raw_store")}/{filename}.pkl', "wb") as pkl:
                pickle.dump(raw_res_store, pkl)