# %%
import warnings
try:
    from pandas.errors import SettingWithCopyWarning
    warnings.simplefilter(action='ignore', category=SettingWithCopyWarning)
except:
    pass
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
warnings.simplefilter(action='ignore', category=RuntimeWarning)

import os
import pandas as pd
from glob import glob
from tqdm import tqdm
import pickle
import sys
import itertools
import argparse

# sys.path.insert(0, os.getcwd())
sys.path.insert(0, "/group/glastonbury/soumick/MyCodes/GitLab/tricorder")

import numpy as np
from scipy import stats
from scipy.stats import norm
from statsmodels.stats.power import tt_ind_solve_power
from sklearn.metrics import roc_auc_score, f1_score

from skimage.filters import threshold_otsu

import pyreadr
from rds2py import read_rds

from openpyxl.styles import Font, Alignment, PatternFill

import matplotlib.pyplot as plt
import seaborn as sns

from lifelines import KaplanMeierFitter, NelsonAalenFitter

from utils.stats.delong import delong_roc_test
from utils.stats.power import simulate_power_wilcoxon
from utils.python_utils import recursive_defaultdict, transform_text

from analyses.disease_mappings import *

import matplotlib.font_manager as fm
font_path = '/project/ukbblatent/soumick/fonts/Helvetica/Helvetica.ttf'  
font_prop = fm.FontProperties(fname=font_path, size=7)
fm.fontManager.addfont(font_path)
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = [font_prop.get_name()]
plt.rcParams['pdf.fonttype'] = 42 #Type 3 fonts (TrueType fonts), which are embedded in the PDF and kept as text rather than being converted to outlines.

# sns.set_style("whitegrid")

# for the prevalence plots, defination of the percentiles
percentiles = [5, 10, 20, 80, 90, 95]
x_labels = ['Bottom 5%', 'Bottom 10%', 'Bottom 20%', 'Top 20%', 'Top 10%', 'Top 5%']

# for the prevalence plots, defination of the colours and markers
# colours = ['salmon', 'turquoise', 'mediumpurple', 'sandybrown', 'lightgreen', 'palevioletred']
colours = ['#FF4D6FFF', '#579EA4FF', '#86AD34FF', '#5D7298FF', '#7E1A2FFF', '#C8350DFF']
markers = ['o', 's', 'D', '^', 'v', 'P']

# %% command-line arguments

def getARGSParser():
    parser = argparse.ArgumentParser(description='MultiPRS Script')
    # parser.add_argument('--output_root', type=str, help='Path to the multi-PRS analyses root directory', default="/group/glastonbury/soumick/PRS/LDPred2/F20208v3_DiffAE/nonDisc/newcovsets_V0v2/4paper_caucasian_king0p0625_grouped/panCohortV2_auto_lw_gw_10kIT_kingB4ldpred2")
    # parser.add_argument('--output_root', type=str, help='Path to the multi-PRS analyses root directory', default="/group/glastonbury/soumick/PRS/LDPred2/F20208v3_DiffAE/nonDisc/newcovsets_V0v2/4paper_caucasian_king0p0625_grouped/panCohortV2_NyDiag_auto_lw_gw_10kIT_kingB4ldpred2")
    parser.add_argument('--output_root', type=str, help='Path to the multi-PRS analyses root directory', default="/group/glastonbury/soumick/PRS/LDPred2/F20208v3_DiffAE/nonDisc/newcovsets_V0v2/4paper_caucasian_king0p0625_grouped/panCohortV2_1to10yProg_auto_lw_gw_10kIT_kingB4ldpred2")
    parser.add_argument('--disease_root', type=str, help='path to the root directory where the disease cohort fils are storred', default="/project/ukbblatent/clinicaldata/binary_disease_cohorts/F20208v3_nonDiscov/caucasian_king0p0625_grouped/newcovsets/V0v2")

    parser.add_argument('--rds_pres_prefix', type=str, help='Prefix (full path) to the PRS file', default="/group/glastonbury/soumick/PRS/LDPred2/F20208v3_DiffAE/nonDisc/run_ext_basic_king0p0625_lw_gw_indep_FiltMAF_")
    parser.add_argument('--rds_pres_suffix', type=str, help='Suffix after the pheno name in the RDS file name', default=".fullDS.auto.mod.LDPred2.rds")
    parser.add_argument('--rds_tag_prs', type=str, help='tag PRS present in the rds file name', default="auto.mod")
    parser.add_argument('--tag_data', type=str, help='tag PRS model', default="resNdata.basic")
    parser.add_argument('--tag_prs', type=str, help='tag PRS inside the rds file', default="pred_auto")

    parser.add_argument('--ext_covar', type=str, help='tag PRS inside the rds file', default="/group/glastonbury/soumick/PRS/inputs/F20208v3_DiffAE_select_latents_r80_discov_INF30/covars/nonDisc_caucasian_king0p0625_V0.tsv")
    
    parser.add_argument('--reprocess_raw_diseases', action=argparse.BooleanOptionalAction, default=False, help='Whether to re-process the raw_diseaes, even if the file already exists in the current output_root')
    parser.add_argument('--raw_disease_path', type=str, help='Path to the raw/unprocessed disease file (not disease cohorts)', default="/project/ukbblatent/clinicaldata/merge_SR_HI_GP_v4_allUKB_&_HEALTHY.csv")
    parser.add_argument('--raw_baseline_path', type=str, help='Path to the raw/unprocessed baseline file (raw output of UKBPuller)', default="/project/ukbblatent/clinicaldata/v1.1.0_seventh_basket/baseline_MD_27_10_2023_13_10_05.tsv")
    parser.add_argument('--raw_centre_info_path', type=str, help='Path to the raw/unprocessed assessment centre info file (raw output of UKBPuller)', default="/project/ukbblatent/clinicaldata/v1.1.0_seventh_basket/assessmentCentre_82779_MD_13_06_2024_12_18_59.tsv")

    parser.add_argument('--drop_dis', type=str, help='Coma-seperated list of diseaes to drop', default="extended arrythmias atrial_ventricular,atrial fibrillation_flutter,arrythmias atrial_ventricular,mitral problems,extended miscellaneous,miscellaneous,aortic valve disorders,pericardial problem,cardiac arrest")
    parser.add_argument('--drop_comparisons', action=argparse.BooleanOptionalAction, default=True, help='Whether to drop comparisons if they are not marked as 1 in the list of comparisons-tags')
    
    parser.add_argument('--is_pancohort', action=argparse.BooleanOptionalAction, default=True, help='Whether to run/plot the pancohort probability predictions')
    parser.add_argument('--save_plots', action=argparse.BooleanOptionalAction, default=True, help='Whether to save the plots')
    parser.add_argument('--save_format', type=str, default="pdf", help='Whether to save the plots')
    parser.add_argument('--sex_stratified', action=argparse.BooleanOptionalAction, default=True, help='Whether to perform in sex-stratified manner as well')

    parser.add_argument('--obtain_summary', action=argparse.BooleanOptionalAction, default=False, help='Whether to obtain summary scores and perform statistical stats. If set to False, it will be presumed that they are already performed and the required files are present')
    parser.add_argument('--obtain_pairwise_improvements', action=argparse.BooleanOptionalAction, default=False, help='Whether to print the pairwise improvements')
    
    parser.add_argument('--split_top_dis_half', action=argparse.BooleanOptionalAction, default=False, help='Whether to split the diseases in half, as top and bottom diseases')
    parser.add_argument('--sort_mode_top_dis', type=str, default="0", help='0: score-based, 1: alphabetically')
    
    parser.add_argument('--plot_box_AUC', action=argparse.BooleanOptionalAction, default=False, help='Whether to plot the box plots for AUC')
    parser.add_argument('--plot_box_F1', action=argparse.BooleanOptionalAction, default=False, help='Whether to plot the box plots for F1')
    parser.add_argument('--plot_box_logOR', action=argparse.BooleanOptionalAction, default=False, help='Whether to plot the box plots for logOR [If pancohort, both pancohort and disease cohort will be plotted]')
    parser.add_argument('--plot_prevalence_prs_pancohort', action=argparse.BooleanOptionalAction, default=True, help='Whether to plot the prevalence vs PRS scores (best performing single PRS) plots for the pancohort')
    parser.add_argument('--plot_prevalence_prob_pancohort', action=argparse.BooleanOptionalAction, default=True, help='Whether to plot the prevalence vs probability plots for the pancohort')

    parser.add_argument('--plot_cum_disease_burden', action=argparse.BooleanOptionalAction, default=False, help='[Disease plot] Whether to plot the cumulative disease burden plots')
    parser.add_argument('--plot_cum_hazard', action=argparse.BooleanOptionalAction, default=False, help='[Disease plot] Whether to plot the cumulative hazard plots')
    parser.add_argument('--plot_KM_survival', action=argparse.BooleanOptionalAction, default=False, help='[Disease plot] Whether to plot the Kaplan-Meier survival plots')
    parser.add_argument('--dis_plots_mod', default="GLM,singlePRSCovar", help='Which predictive model to use for the disease plots. Leave it blank if not desired')
    parser.add_argument('--dis_plots_rawPRS', action=argparse.BooleanOptionalAction, default=True, help='[Disease plot] Whether to plot disease plots using raw PRS scores (Latent selected using the best performing single PRS for that disease)')
    parser.add_argument('--dis_plots_cutoff_date', default="2023-10-31", help='The cutoff date to be used (ideally, the download date) for the disease plots - to filter data with strange dates')
    parser.add_argument('--dis_plots_upto_Nyear', default=10, type=int, help='How many years to consider in the disease plots')    
    
    parser.add_argument('--save_raw', action=argparse.BooleanOptionalAction, default=True, help='[Disease and prevalence plots] Whether to store the raw values used to create the plots')

    parser.add_argument('--box_limits_AUC', type=str, default="0.55,0.80,0.05", help='Upper bound, lower bound, and distance for the y-axis of the AUC plots [Leave blank for using default]')
    parser.add_argument('--box_limits_F1', type=str, default="0.50,0.75,0.05", help='Upper bound, lower bound, and distance for the y-axis of the AUC plots [Leave blank for using default]')
    parser.add_argument('--box_limits_logOR', type=str, default="0.50,2.25,0.25", help='Upper bound, lower bound, and distance for the y-axis of the AUC plots [Leave blank for using default]')
    parser.add_argument('--box_limits_logOR_pancohort', type=str, default="0.50,2.00,0.25", help='Upper bound, lower bound, and distance for the y-axis of the AUC plots [Leave blank for using default]')
    
    parser.add_argument('--colour_box', type=str, default="#A4203D,#D87412,#5A6F94", help='List of colours to be used with the box plots [Leave blank for using default]')
    parser.add_argument('--colour_transparency_box', type=float, default=1, help='The colour_transparency percentage between 0 and 1 for the colours of the box plots [Set it to 0 or 1 to ignore]')
    parser.add_argument('--colour_prevalence', type=str, default="#FD4C6E,#80A531,#559AA0", help='List of colours to be used with the prevalence plots [Leave blank for using default]')
    parser.add_argument('--colour_displots', type=str, default="#C7350D,#F6BD01,#81B18C", help='List of colours to be used with the disease plots [Leave blank for using default]')

    return parser

# %% Decorators
def plots_and_results_by_sex(func):
    def wrapper(*args, **kwargs):
        sex = ['Both', 'Female', 'Male']
        results = []
        for s in sex:
            new_kwargs = kwargs.copy()
            if s == 'Female' and 'IIDs_female' in kwargs:
                new_kwargs['IIDs'] = kwargs['IIDs_female']
            elif s == 'Male' and 'IIDs_male' in kwargs:
                new_kwargs['IIDs'] = kwargs['IIDs_male']
            elif 'IIDs_female' in kwargs and 'IIDs_female' in kwargs:
                new_kwargs['IIDs'] = kwargs['IIDs_female'] + kwargs['IIDs_male']
            if 'IIDs' in new_kwargs:
                del new_kwargs['IIDs_female']
                del new_kwargs['IIDs_male']
            new_kwargs['sex'] = s
            result = func(*args, **new_kwargs)
            results.append(result)
        return results
    return wrapper

# %% Support functions for the processing of the raw results and compute the stats

def get_stats(score_m0, score_m1):
    try:
        t_stat, p_value_t_test = stats.ttest_rel(score_m0, score_m1)
        print("Paired t-test - p-value:", p_value_t_test, f"({'Sig' if p_value_t_test < 0.05 else 'Not Sig'})")

        effect_size = (np.mean(score_m0-score_m1) / np.std(score_m0-score_m1))

        power = tt_ind_solve_power(effect_size=effect_size, nobs1=5, alpha=0.05, ratio=1, alternative='two-sided')
        print("Statistical Power (Paired t-test):", power)

        stat, p_value_wilcoxon = stats.wilcoxon(score_m0, score_m1)
        print("Wilcoxon Signed-Rank Test - p-value:", p_value_wilcoxon, f"({'Sig' if p_value_wilcoxon < 0.05 else 'Not Sig'})")
        
        estimated_power = simulate_power_wilcoxon(len(score_m0), effect_size)
        print("Estimated Power:", estimated_power)

        return p_value_t_test, power, p_value_wilcoxon, estimated_power
    except:
        return np.nan, np.nan, np.nan, np.nan

def get_sig(method0, method1, comboDF=None, pancohort=False):
    if type(method0) == str:
        method0 = comboDF[comboDF['model'] == method0]
    if type(method1) == str:
        method1 = comboDF[comboDF['model'] == method1]

    method0 = method0.sort_values('fold')
    method1 = method1.sort_values('fold')

    print("AUC:")
    p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC = get_stats(method0['AUC_test'].values, method1['AUC_test'].values)

    print("logOR:")
    p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR = get_stats(method0['logOR_test'].values, method1['logOR_test'].values)

    if pancohort:
        print("logOR (pancohort):")
        p_value_t_test_logOR_pancohort, power_t_test_logOR_pancohort, p_value_wilcoxon_logOR_pancohort, power_wilcoxon_logOR_pancohort = get_stats(method0['logOR_test_pancohort'].values, method1['logOR_test_pancohort'].values)

        return p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC, p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR, p_value_t_test_logOR_pancohort, power_t_test_logOR_pancohort, p_value_wilcoxon_logOR_pancohort, power_wilcoxon_logOR_pancohort
    else:
        return p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC, p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR

def getSummaryDF(df, pancohort=False):
    cols = ['AUC_test', 'logOR_test', 'logOR_test_pancohort', 'F1_test'] if pancohort else ['AUC_test', 'logOR_test', 'F1_test']
    group_cols = ['tag', 'Sex'] if "Sex" in df else ['tag']
    median_df = df.groupby(group_cols)[cols].median()   
    iqr_df = (df.groupby(group_cols)[cols].quantile(0.75) - df.groupby(group_cols)[cols].quantile(0.25))
    merged_df = median_df.join(iqr_df, lsuffix='_median', rsuffix='_iqr')
    merged_df['AUC_test'] = merged_df.apply(lambda row: f"{row['AUC_test_median']:.4f} ± {row['AUC_test_iqr']:.4f}", axis=1)
    merged_df['logOR_test'] = merged_df.apply(lambda row: f"{row['logOR_test_median']:.4f} ± {row['logOR_test_iqr']:.4f}", axis=1)
    merged_df['F1_test'] = merged_df.apply(lambda row: f"{row['F1_test_median']:.4f} ± {row['F1_test_iqr']:.4f}", axis=1)
    if pancohort:
        merged_df['logOR_test_pancohort'] = merged_df.apply(lambda row: f"{row['logOR_test_pancohort_median']:.4f} ± {row['logOR_test_pancohort_iqr']:.4f}", axis=1)
    return merged_df[cols].reset_index()

# %% Support functions to creating summary tables

def extract_median_iqr(value):
    median, iqr = value.split(' ± ')
    return float(median), float(iqr)

def determine_winner(row, col0, col1):
    median0, iqr0 = extract_median_iqr(row[col0])
    median1, iqr1 = extract_median_iqr(row[col1])    
    if median0 > median1:
        return row['Method0']
    elif median0 < median1:
        return row['Method1']
    else:
        if iqr0 == iqr1:
            return 'Tie'
        return row['Method0'] if iqr0 < iqr1 else row['Method1']

def combineDFs(statsig, sumres):
    merged_df = statsig.merge(sumres, left_on=['Method0', 'Sex', 'Disease'], right_on=['tag', 'Sex', 'Disease'], suffixes=('', '_Method0'))
    merged_df = merged_df.merge(sumres, left_on=['Method1', 'Sex', 'Disease'], right_on=['tag', 'Sex', 'Disease'], suffixes=('_Method0', '_Method1'))

    combined_df = merged_df[['Disease', 'Sex', 'Comparison', 'Method0', 'Method1', 'p_value_DeLong_AUC', 
                            'AUC_test_Method0', 'logOR_test_Method0', 'logOR_test_pancohort_Method0', 'F1_test_Method0', 
                            'AUC_test_Method1', 'logOR_test_Method1', 'logOR_test_pancohort_Method1', 'F1_test_Method1', ]]

    combined_df['Significant'] = combined_df['p_value_DeLong_AUC'] < 0.05
    combined_df['AUCWinner'] = combined_df.apply(determine_winner, axis=1, args=('AUC_test_Method0', 'AUC_test_Method1'))
    combined_df['logORWinner'] = combined_df.apply(determine_winner, axis=1, args=('logOR_test_Method0', 'logOR_test_Method1'))
    combined_df['logORPancohortWinner'] = combined_df.apply(determine_winner, axis=1, args=('logOR_test_pancohort_Method0', 'logOR_test_pancohort_Method1'))
    combined_df['F1Winner'] = combined_df.apply(determine_winner, axis=1, args=('F1_test_Method0', 'F1_test_Method1'))

    return combined_df

def summary_table(combined_df, ignore_methods=[], diseases=[]):
    def determine_result(row, col):
        winner = row[col]
        if winner == 'Tie':
            return 'Tie'
        if row['Significant']:
            if winner == row['Method0']:
                return f'Significantly Better for {row["Method0"]}'
            else:
                return f'Significantly Better for {row["Method1"]}'
        else:
            if winner == row['Method0']:
                return f'Insignificantly Better for {row["Method0"]}'
            else:
                return f'Insignificantly Better for {row["Method1"]}'

    combined_df['AUCResult'] = combined_df.apply(determine_result, axis=1, args=('AUCWinner',))
    combined_df['logORResult'] = combined_df.apply(determine_result, axis=1, args=('logORWinner',))
    combined_df['logORPancohortResult'] = combined_df.apply(determine_result, axis=1, args=('logORPancohortWinner',))
    combined_df['F1Result'] = combined_df.apply(determine_result, axis=1, args=('F1Winner',))

    if bool(diseases):
        combined_df = combined_df[combined_df['Disease'].isin(diseases)]

    summary = []

    for sex in combined_df['Sex'].unique():
        for comparison in combined_df['Comparison'].unique():
            sub_df = combined_df[(combined_df['Sex'] == sex) & (combined_df['Comparison'] == comparison)]
            auc_counts = sub_df['AUCResult'].value_counts().to_dict()
            logor_counts = sub_df['logORResult'].value_counts().to_dict()
            logor_pancohort_counts = sub_df['logORPancohortResult'].value_counts().to_dict()
            f1_counts = sub_df['F1Result'].value_counts().to_dict()

            if sub_df['Method0'].iloc[0] in ignore_methods or sub_df['Method1'].iloc[0] in ignore_methods:
                continue
            
            summary.append({
                'Sex': sex,
                'Comparison': comparison,
                'Method0': sub_df['Method0'].iloc[0],
                'Method1': sub_df['Method1'].iloc[0],
                'AUC_Significantly_Better_for_Method0': auc_counts.get(f'Significantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'AUC_Significantly_Better_for_Method1': auc_counts.get(f'Significantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'AUC_Insignificantly_Better_for_Method0': auc_counts.get(f'Insignificantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'AUC_Insignificantly_Better_for_Method1': auc_counts.get(f'Insignificantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'logOR_Significantly_Better_for_Method0': logor_counts.get(f'Significantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'logOR_Significantly_Better_for_Method1': logor_counts.get(f'Significantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'logOR_Insignificantly_Better_for_Method0': logor_counts.get(f'Insignificantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'logOR_Insignificantly_Better_for_Method1': logor_counts.get(f'Insignificantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'logOR_Pancohort_Significantly_Better_for_Method0': logor_pancohort_counts.get(f'Significantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'logOR_Pancohort_Significantly_Better_for_Method1': logor_pancohort_counts.get(f'Significantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'logOR_Pancohort_Insignificantly_Better_for_Method0': logor_pancohort_counts.get(f'Insignificantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'logOR_Pancohort_Insignificantly_Better_for_Method1': logor_pancohort_counts.get(f'Insignificantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'F1_Significantly_Better_for_Method0': f1_counts.get(f'Significantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'F1_Significantly_Better_for_Method1': f1_counts.get(f'Significantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'F1_Insignificantly_Better_for_Method0': f1_counts.get(f'Insignificantly Better for {sub_df["Method0"].iloc[0]}', 0),
                'F1_Insignificantly_Better_for_Method1': f1_counts.get(f'Insignificantly Better for {sub_df["Method1"].iloc[0]}', 0),
                'Number_of_Diseases': len(sub_df)
            })

    summary_df = pd.DataFrame(summary)
    return summary_df

def create_excel_summary(summary_df, metric_col, output_path):
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    for sex in summary_df['Sex'].unique():
        filtered_df = summary_df[summary_df['Sex'] == sex]
        
        summary_data = []
        for _, row in filtered_df.iterrows():
            summary_data.append([
                row['Method0'],
                row['Method1'],
                row[f'{metric_col}_Significantly_Better_for_Method0'],
                row[f'{metric_col}_Insignificantly_Better_for_Method0'],
                row[f'{metric_col}_Significantly_Better_for_Method1'],
                row[f'{metric_col}_Insignificantly_Better_for_Method1']
            ])
        
        summary_formatted_df = pd.DataFrame(summary_data, columns=[
            'Method0', 'Method1', 'Sig_Better_Method0', 'Insig_Better_Method0', 'Sig_Better_Method1', 'Insig_Better_Method1'
        ])
        summary_formatted_df.to_excel(writer, index=False, sheet_name=sex, startrow=2, header=False)

        workbook = writer.book
        worksheet = writer.sheets[sex]

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        worksheet.merge_cells('A1:B1')
        worksheet['A1'] = 'Methods'
        worksheet['A1'].font = header_font
        worksheet['A1'].alignment = header_alignment
        worksheet['A1'].fill = header_fill

        worksheet.merge_cells('C1:D1')
        worksheet['C1'] = 'better Method0'
        worksheet['C1'].font = header_font
        worksheet['C1'].alignment = header_alignment
        worksheet['C1'].fill = header_fill

        worksheet.merge_cells('E1:F1')
        worksheet['E1'] = 'better Method1'
        worksheet['E1'].font = header_font
        worksheet['E1'].alignment = header_alignment
        worksheet['E1'].fill = header_fill

        sub_headers = ['Method0', 'Method1', 'Sig', 'Insig', 'Sig', 'Insig']
        for col_num, sub_header in enumerate(sub_headers, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.value = sub_header
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

    writer.save()

@plots_and_results_by_sex
def compareSummaryPairs(df, all_combinations, disease, sex, pancohort=False):
    collect = []
    for combination in all_combinations:
        comparison1, comparison2 = combination
        print(f"(Sex: {sex}) {comparison1[2]} [vs] {comparison2[2]}-----")

        try:
            df_sex = df[df.Sex == sex]
            df_method1 = df_sex[df_sex.apply(lambda row: (row['method'], row['res_type']) == (comparison1[0], comparison1[1]), axis=1)]
            df_method2 = df_sex[df_sex.apply(lambda row: (row['method'], row['res_type']) == (comparison2[0], comparison2[1]), axis=1)]
            
            if pancohort:
                p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC, p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR, p_value_t_test_logOR_pancohort, power_t_test_logOR_pancohort, p_value_wilcoxon_logOR_pancohort, power_wilcoxon_logOR_pancohort = get_sig(df_method1, df_method2, pancohort=True)
            else:
                p_value_t_test_AUC, power_t_test_AUC, p_value_wilcoxon_AUC, power_wilcoxon_AUC, p_value_t_test_logOR, power_t_test_logOR, p_value_wilcoxon_logOR, power_wilcoxon_logOR = get_sig(df_method1, df_method2)

            res = {
                    "Disease": disease,
                    "Sex": sex,
                    "Comparison": f"{comparison1[2]} [vs] {comparison2[2]}",
                    "Method0": comparison1[2],
                    "Method1": comparison2[2],
                    "p_value_t_test_AUC": p_value_t_test_AUC,
                    "power_t_test_AUC": power_t_test_AUC,
                    "p_value_wilcoxon_AUC": p_value_wilcoxon_AUC,
                    "power_wilcoxon_AUC": power_wilcoxon_AUC,
                    "p_value_t_test_logOR": p_value_t_test_logOR,
                    "power_t_test_logOR": power_t_test_logOR,
                    "p_value_wilcoxon_logOR": p_value_wilcoxon_logOR,
                    "power_wilcoxon_logOR": power_wilcoxon_logOR
                }
            if pancohort:
                res.update({
                    "p_value_t_test_logOR_pancohort": p_value_t_test_logOR_pancohort,
                    "power_t_test_logOR_pancohort": power_t_test_logOR_pancohort,
                    "p_value_wilcoxon_logOR_pancohort": p_value_wilcoxon_logOR_pancohort,
                    "power_wilcoxon_logOR_pancohort": power_wilcoxon_logOR_pancohort
                })
            collect.append(res)
        except:
            pass
        
    return pd.DataFrame(collect)

def getIndividualProbs(res_method, res_type, IIDs, get_pancohort=False, resDF=None):
    if "single" in res_type.lower():
        if resDF is not None:
            latent = resDF.iloc[0].Latent
            res_method = res_method.loc[latent]
        else:
            res_method = res_method.iloc[res_method['AUC_test'].argmax()]
    res = pd.DataFrame(res_method['pred_probs_test_pancohort']) if get_pancohort else pd.DataFrame(res_method['pred_probs_test'])
    res = res[res.index.isin(IIDs)]
    return res, res_method

@plots_and_results_by_sex
def compareAUCPairs(res, all_combinations, disease, diseaseDF, IIDs, sex, resDF):
    collect = []
    for combination in all_combinations:
        comparison1, comparison2 = combination
        print(f"(Sex: {sex}) {comparison1[2]} [vs] {comparison2[2]} -----")
        
        try:
            collect_fold = []
            for fold in res.keys():
                pred_prob_m0, _ = getIndividualProbs(res[fold][comparison1[0]][comparison1[1]], comparison1[1], IIDs=IIDs, resDF=resDF[(resDF.tag == comparison1[2]) & (resDF.Sex == sex) & (resDF.fold == fold)])
                pred_prob_m1, _ = getIndividualProbs(res[fold][comparison2[0]][comparison2[1]], comparison2[1], IIDs=IIDs, resDF=resDF[(resDF.tag == comparison2[2]) & (resDF.Sex == sex) & (resDF.fold == fold)])
                pred_prob = pred_prob_m0.join(pred_prob_m1, lsuffix="_m0", rsuffix="_m1")
                combined = pred_prob.join(diseaseDF[['BinCAT_Disease']])
                delongP = delong_roc_test(combined['BinCAT_Disease'], combined['predicted_m0'], combined['predicted_m1'])
                print(f"DeLong's p-value [{fold}]:", delongP, f"({'Sig' if delongP < 0.05 else 'Not Sig'})")
                collect_fold.append(combined)
            collect_fold = pd.concat(collect_fold)
            delongP = delong_roc_test(collect_fold['BinCAT_Disease'], collect_fold['predicted_m0'], collect_fold['predicted_m1'])
            print(f"DeLong's p-value [Across Folds]:", delongP, f"({'Sig' if delongP < 0.05 else 'Not Sig'})")

            collect.append(
                {
                    "Disease": disease,
                    "Sex": sex,
                    "Comparison": f"{comparison1[2]} [vs] {comparison2[2]}",
                    "Method0": comparison1[2],
                    "Method1": comparison2[2],
                    "p_value_DeLong_AUC": delongP
                }
            )
        except:
            pass
    return pd.DataFrame(collect)


# %% Support functions related to the plotting

def prevalence_and_ci(pred_probs, true_labels, percentile, confidence=0.95):
    if percentile < 50:
        group = pred_probs <= np.percentile(pred_probs, percentile)
    else:
        group = pred_probs >= np.percentile(pred_probs, percentile)

    prevalence = true_labels[group].mean()
    n = len(true_labels[group])
    se = np.sqrt(prevalence * (1 - prevalence) / n)  
    ci_half_width = se * norm.ppf((1 + confidence) / 2)  
    return prevalence, prevalence - ci_half_width, prevalence + ci_half_width, n

def z_test(p_hat, p_0, n):
    se = np.sqrt(p_0 * (1 - p_0) / n)  
    z = (p_hat - p_0) / se  
    p_value = 2 * (1 - norm.cdf(abs(z)))  # Two-tailed test
    return p_value

def is_significant(p_hat, p_0, n, alpha=0.05):
    p_value = z_test(p_hat, p_0, n)
    return p_value < alpha

def generate_sequence_plotdelta(N):
    sequence = []
    if N == 1:
        sequence.append(0)
    else:
        for i in range(-(N//2), N//2 + 1):
            sequence.append(i * 0.1)
        if N % 2 == 0:
            sequence.remove(0)
    return sequence

def plot_prevalence(df, res_keys=['Pred_Covariates', 'Pred_PRS'], res_labels=['Covariates', 'PRS'], res_colours=['orange', 'blue'], res_fmts=['o', 's'], col_disease='BinCAT_Disease', plot_title='prova!', individual_prev_line=False, save_path="", flip_PRS_sign=False, save_raw=False):
    #supply Pred_Covariates, Pred_PRS and BinCAT_Disease as columns of the DataFrame 

    if save_raw:
        prevalence_data = {'Female': [], 'Both': [], 'Male': []}
        ci_lower_data = {'Female': [], 'Both': [], 'Male': []}
        ci_upper_data = {'Female': [], 'Both': [], 'Male': []}
        percentile_labels = ['Bottom 5%', 'Bottom 10%', 'Bottom 20%', 'Top 20%', 'Top 10%', 'Top 5%']

    plotdelta = generate_sequence_plotdelta(len(res_keys))
    x_ticks = np.arange(len(x_labels))

    overall_prevalence = df[col_disease].mean()
    mid_percentile = x_ticks.mean()

    plt.figure(constrained_layout=True)
    fig, ax = plt.subplots(figsize=(10, 6))

    for i in range(len(res_keys)):
        df_curr = df[[res_keys[i], col_disease]].dropna() #this will handle the sex-stratified case
        results = [prevalence_and_ci(-df_curr[res_keys[i]] if flip_PRS_sign else df_curr[res_keys[i]], df_curr[col_disease], p) for p in percentiles]
        bottom, bottom_cis_lower, bottom_cis_upper, bottom_n = zip(*results[:3])
        top, top_cis_lower, top_cis_upper, top_n = zip(*results[3:])        

        if save_raw:
            prevalence_data[res_labels[i]].extend(bottom + top)
            ci_lower_data[res_labels[i]].extend(bottom_cis_lower + top_cis_lower)
            ci_upper_data[res_labels[i]].extend(bottom_cis_upper + top_cis_upper)
        
        ax.errorbar(np.arange(len(bottom)) + plotdelta[i], bottom, yerr=[np.subtract(bottom, bottom_cis_lower), np.subtract(bottom_cis_upper, bottom)], fmt=res_fmts[i], color=res_colours[i], label=res_labels[i], capsize=5)
        ax.errorbar(np.arange(len(top)) + 3 + plotdelta[i], top, yerr=[np.subtract(top, top_cis_lower), np.subtract(top_cis_upper, top)], fmt=res_fmts[i], color=res_colours[i], capsize=5)

        if individual_prev_line:
            ax.axhline(y=df_curr[col_disease].mean(), color=res_colours[i], linestyle='--')

    if not individual_prev_line:
        ax.axhline(y=overall_prevalence, color='gray', linestyle='--')
    ax.axvline(x=mid_percentile, color='gray', linestyle='dotted')

    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    ax.set_xticks(x_ticks)
    ax.set_xticklabels(x_labels, fontsize=13)
    ax.set_xlabel('Percentile', fontsize=15)
    ax.set_ylabel('Prevalence', fontsize=15)
    ax.set_title(plot_title, fontsize=16)
    ax.legend(fontsize=13, title_fontsize=15)

    ax.grid(False)

    if bool(save_path):
        plt.savefig(save_path, dpi=300)
    else:
        plt.show()

    if save_raw:
        return {"prevalence_data": prevalence_data, "ci_lower_data": ci_lower_data, "ci_upper_data": ci_upper_data, "percentile_labels": percentile_labels}

@plots_and_results_by_sex
def plot_box(df, sex, x='Disease', y='AUC_test', y_tag='AUC', legends_tag="Method", hue='tag', hue_order=None, palette="Set2", colour_transparency=0.6, y_tick_limits="", save_path_noext="", save_format="pdf"):
    df_sex = df[df.Sex == sex]

    plt.figure(figsize=(14, 8), constrained_layout=True)

    ax = sns.boxplot(data=df_sex, x=x, y=y, hue=hue, hue_order=hue_order, palette=palette, saturation=1, linewidth=1)
    if colour_transparency not in [0,1]:
        for patch in ax.patches:
            r, g, b, a = patch.get_facecolor()
            patch.set_facecolor((r, g, b, colour_transparency))

    # ax.spines['left'].set_visible(False)
    # ax.spines['bottom'].set_visible(False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.title(f'{y_tag} by {x} and {legends_tag} (Sex: {sex})', fontsize=19)
    plt.xlabel(x, fontsize=18)
    plt.ylabel(y_tag, fontsize=18)
    plt.legend(title=legends_tag, fontsize=16, title_fontsize=18)
    plt.xticks(rotation=45, fontsize=16)
    plt.grid(False)
    # plt.grid(True, which='both', color='grey', linestyle='--', linewidth=0.5, alpha=0.4)

    if bool(y_tick_limits):
        y_tick_limits = [float(t) for t in y_tick_limits.split(",")]
        plt.ylim(y_tick_limits[0]-0.02, y_tick_limits[1]+0.02)
        plt.yticks(np.arange(y_tick_limits[0], y_tick_limits[1]+0.02, y_tick_limits[2]), fontsize=16)
    else:
        plt.yticks(fontsize=16)

    if bool(save_path_noext):
        plt.savefig(f"{save_path_noext}_{sex}.{save_format}", dpi=300)
    else:
        plt.show()

@plots_and_results_by_sex
def plot_prevalence_prob_pancohort(res, disease, diseaseDF, collect_df, sex, IIDs, comparisons, fullcohort=True, save_format="pdf"):
    best_fold = collect_df[(collect_df.Disease == disease) & (collect_df.res_type == "singlePRSCovar") & (collect_df.Sex == sex)].sort_values('AUC_test', ascending=False).iloc[0]

    probs = []
    tags = []
    for c in comparisons:
        try:
            resDF = collect_df[(collect_df.Disease == disease) & (collect_df.tag == c[2]) & (collect_df.Sex == sex) & (collect_df.fold == best_fold.fold)]
            pred_prob, _ = getIndividualProbs(res[best_fold.fold][c[0]][c[1]], c[1], IIDs=IIDs, resDF=resDF, get_pancohort=True)
            pred_prob.rename({'predicted': c[1]}, axis=1, inplace=True)
            probs.append(pred_prob)
            tags.append(c[2])
        except:
            pass
    probs = pd.concat(probs, axis=1)

    combined = probs.join(diseaseDF[['BinCAT_Disease']])

    if fullcohort:
        combined.fillna(0, inplace=True)
    else:
        combined
    combined["BinCAT_Disease"] = combined["BinCAT_Disease"].astype(int)

    res_cols = [c for c in combined.columns if c!="BinCAT_Disease"]

    if args.save_plots:
        pth = f"{args.output_root}/plots/prevalence_prob_pancohort/bestfold_{d.replace(' ', '-')}_{sex}.{save_format}"
    else:
        pth = ""
    return plot_prevalence(combined, res_keys=res_cols, res_labels=tags, res_colours=args.colour_prevalence[:len(tags)], res_fmts=markers[:len(tags)], col_disease='BinCAT_Disease', plot_title=f'Risk scores (Full cohort [Sex: {sex}], excluding discovery and PRS training sets): {d.replace("-", " ")}', save_path=pth)

@plots_and_results_by_sex
def plot_dis_plots(res, disease, diseaseDF, collect_df, sex, IIDs, args, model_type="GLM", model="singlePRSCovar", PRS_col="predicted", ending_year = 10, use_raw_PRS=False, res_colours=['#FF4D6FFF', '#579EA4FF', '#86AD34FF'], save_format="pdf", save_raw=False):
    if use_raw_PRS:
        model_type = "GLM"
        model = "singlePRS"
        tag = "rawPRS_"
    else:
        tag = ""

    if save_raw:
        raw_store = {"Type": "bestfold" if tag=="" else "rawPRS", "Disease": disease, "Sex": sex}

    resDF = collect_df[(collect_df.Disease == disease) & (collect_df.method == model_type) & (collect_df.res_type == model) & (collect_df.Sex == sex)].sort_values('AUC_test', ascending=False)
    if use_raw_PRS:
        pred_prob = getBestPRS(rds_pres_prefix=args.rds_pres_prefix, rds_pres_suffix=args.rds_pres_suffix, rds_tag_prs=args.rds_tag_prs, tag_data=args.tag_data, tag_prs=args.tag_prs, IIDs=IIDs, resDF=resDF)
    else:
        pred_prob, _ = getIndividualProbs(res[resDF.iloc[0].fold][model_type][model], model, IIDs=IIDs, resDF=resDF, get_pancohort=True)
    prob_days = pred_prob.merge(diseaseDF[diseaseDF.summary == disease.lower()]['DiseaseAfter'], left_index=True, right_index=True)
    prob_control = pred_prob.merge(diseaseDF[diseaseDF.summary == "healthy"]['DiseaseAfter'], left_index=True, right_index=True)

    if args.plot_cum_disease_burden:
        prob_days['PRSQuantile'] = pd.qcut(prob_days[PRS_col], q=3, labels=['Low risk PRS', 'Mid risk PRS', 'Top risk PRS'])
        prob_days['Years'] = prob_days['DiseaseAfter'] / 365.25
        time_periods = np.arange(1, prob_days['Years'].max()+1 if ending_year == -1 else ending_year)  
        quantiles = ['Top risk PRS', 'Mid risk PRS', 'Low risk PRS']

        plot_data = {quantile: [] for quantile in quantiles}

        for year in time_periods:
            for quantile in quantiles:
                count = prob_days[(prob_days['PRSQuantile'] == quantile) & (prob_days['Years'] <= year)].shape[0]
                plot_data[quantile].append(count)
        
        if save_raw:
            raw_store['cum_disease_burden'] = plot_data

        plt.figure(figsize=(6, 6), constrained_layout=True)
        for i, quantile in enumerate(quantiles):
            plt.plot(time_periods, plot_data[quantile], label=quantile, color=res_colours[i])

        plt.xlabel('Years', fontsize=15)
        plt.ylabel('Cumulative number of diagnoses', fontsize=15)
        plt.title(f'{model} for {disease} [Sex: {sex}]', fontsize=16)
        plt.legend(title='PRS Quantiles', fontsize=13, title_fontsize=15)
        plt.xticks(fontsize=13)
        plt.yticks(fontsize=13)
        plt.grid(False)

        ax = plt.gca()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        if args.save_plots:
            plt.savefig(f"{args.output_root}/plots/{tag}cumulative_disease_burden/bestfold_{disease.replace(' ', '-')}_{sex}.{save_format}", dpi=300)
        else:
            plt.show()

    if args.plot_cum_hazard or args.plot_KM_survival:
        prob_days = prob_days[prob_days.DiseaseAfter <= 365*ending_year] #TODO: if we decide to keep it, make it a flag.
        
        prob_incident_control = pd.concat([prob_days[["predicted", "DiseaseAfter"]], prob_control])
        prob_incident_control['event_occurred'] = prob_incident_control['DiseaseAfter'].notna()
        prob_incident_control = prob_incident_control[prob_incident_control['DiseaseAfter'].gt(0) | prob_incident_control['DiseaseAfter'].isna()]
        prob_incident_control['DiseaseAfter'] = prob_incident_control['DiseaseAfter'].fillna(prob_incident_control['DiseaseAfter'].max())
        prob_incident_control['DiseaseAfter'] = prob_incident_control['DiseaseAfter'] / 365.25
        prob_incident_control['PRSQuantile'] = pd.qcut(prob_incident_control[PRS_col], q=3, labels=['Low risk PRS', 'Mid risk PRS', 'Top risk PRS'])

        if args.plot_cum_hazard:
            naf = NelsonAalenFitter()
            if save_raw:
                raw_store['cum_hazard'] = {}

            plt.figure(figsize=(6, 6), constrained_layout=True)
            for i, risk_group in enumerate(['Top risk PRS', 'Mid risk PRS', 'Low risk PRS']):
                mask = prob_incident_control['PRSQuantile'] == risk_group
                naf.fit(durations=prob_incident_control['DiseaseAfter'][mask], event_observed=prob_incident_control['event_occurred'][mask], label=risk_group)
                if save_raw:
                    raw_store['cum_hazard'][risk_group] = naf.cumulative_hazard_
                naf.plot_cumulative_hazard(color=res_colours[i])

            plt.title(f'{model} for {disease} [Sex: {sex}]', fontsize=16)
            plt.xlabel('Years after baseline', fontsize=15)
            plt.ylabel('Cumulative hazard', fontsize=15)
            plt.xlim(0, ending_year)  
            plt.legend(title='PRS risk group', fontsize=13, title_fontsize=15)
            plt.xticks(fontsize=13)
            plt.yticks(fontsize=13)
            plt.grid(False)

            ax = plt.gca()
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            if args.save_plots:
                plt.savefig(f"{args.output_root}/plots/{tag}cumulative_hazard/bestfold_{disease.replace(' ', '-')}_{sex}.{save_format}", dpi=300)
            else:
                plt.show()

        if args.plot_KM_survival:
            kmf = KaplanMeierFitter()
            if save_raw:
                raw_store['KM_survival'] = {}

            plt.figure(figsize=(6, 6), constrained_layout=True)
            for i, risk_group in enumerate(['Top risk PRS', 'Mid risk PRS', 'Low risk PRS']):
                mask = prob_incident_control['PRSQuantile'] == risk_group
                kmf.fit(durations=prob_incident_control['DiseaseAfter'][mask], event_observed=prob_incident_control['event_occurred'][mask], label=risk_group)
                if save_raw:
                    raw_store['KM_survival'][risk_group] = kmf.survival_function_
                kmf.plot(color=res_colours[i])

            plt.title(f'{model} for {disease} [Sex: {sex}]', fontsize=16)
            plt.xlabel('Years after baseline', fontsize=15)
            plt.ylabel('Survival probability', fontsize=15)
            plt.xlim(0, ending_year)  
            plt.legend(title='PRS risk group', fontsize=14, title_fontsize=15)
            plt.xticks(fontsize=13)
            plt.yticks(fontsize=13)
            plt.grid(False)

            ax = plt.gca()
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            if args.save_plots:
                plt.savefig(f"{args.output_root}/plots/{tag}KM_survival/bestfold_{disease.replace(' ', '-')}_{sex}.{save_format}", dpi=300)
            else:
                plt.show()

    if save_raw:
        return raw_store
    else:
        return None
    
# %% Other support functions

def getImprovedDisease(df, tag0, tag1):
        filtered_df = df[df['tag'].isin([tag0, tag1])]
        pivot_df = filtered_df.pivot(index='Disease', columns='tag', values='AUC_test').reset_index()
        result_df = pivot_df[pivot_df[tag0] < pivot_df[tag1]]
        return result_df['Disease'].tolist()

def getBestPRS(rds_pres_prefix, rds_pres_suffix, rds_tag_prs, tag_data, tag_prs, IIDs, res_method=None, resDF=None):
    if resDF is not None:
        best_prs_latent = resDF.iloc[0].Latent
    else:
        best_prs_latent = res_method.iloc[res_method['AUC_test'].argmax()].name
    pth_prs = f"{rds_pres_prefix}{best_prs_latent.split('PRS:')[-1]}{rds_pres_suffix}"

    data = pyreadr.read_r(pth_prs.replace(rds_tag_prs, tag_data))[None]
    prs = data[['IID']]
    r_obj = read_rds(pth_prs)
    prs["predicted"] = r_obj['data'][r_obj['attributes']['names']['data'].index(tag_prs)]['data']
    prs.set_index("IID", inplace=True)
    prs = prs[prs.index.isin(IIDs)]

    return prs

def select_values(row, base_cols, suffixes):
    for suffix in suffixes:
        if all(not pd.isna(row[f"{col}.{suffix}.0"]) for col in base_cols):
            return [row[f"{col}.{suffix}.0"] for col in base_cols] + [suffix]
    return [pd.NA] * len(base_cols)

def get_disease_info(disease_path, raw_baseline_path, info_assessment_centre_path, cutoff_date):
    diseases = pd.read_csv(disease_path, low_memory=False).set_index("eid")

    raw_baseline = pd.read_table(raw_baseline_path).set_index("f.eid")
    raw_baseline = raw_baseline[[c for c in raw_baseline.columns if "f.20116." in c or "f.21001" in c]]
    assessment_centre = pd.read_table(info_assessment_centre_path).set_index("f.eid")
    assessment_centre = assessment_centre[[c for c in assessment_centre.columns if "f.53." in c]]
    df_unprocessed = diseases.merge(raw_baseline, left_index=True, right_index=True).merge(assessment_centre, left_index=True, right_index=True) 
    df_unprocessed.index.names = ['eid']
    df_unprocessed[['f.20116', 'f.21001', 'f.53', 'selected_instance']] = df_unprocessed.apply(lambda row: pd.Series(select_values(row, base_cols=['f.20116', 'f.21001', 'f.53'], suffixes=['0', '1', '2', '3'])), axis=1)
    df_unprocessed = df_unprocessed.drop([c for c in df_unprocessed.columns if "f.20116." in c or "f.21001." in c or "f.53." in c], axis=1) #we will remove now the instance-wise columns
    df_unprocessed['date'] = pd.to_datetime(df_unprocessed['date'], format='%Y-%m-%d', errors='coerce')
    df_unprocessed['f.53'] = pd.to_datetime(df_unprocessed['f.53'], format='%Y-%m-%d', errors='coerce')

    #remove subjects with "future" diseases
    df_unprocessed = df_unprocessed[((df_unprocessed['f.53'].isna()) | (df_unprocessed['f.53'] <= pd.to_datetime(cutoff_date))) & ((df_unprocessed['date'].isna()) | (df_unprocessed['date'] <= pd.to_datetime(cutoff_date)))]

    df_unprocessed['DiseaseAfter'] = (df_unprocessed['date'] - df_unprocessed['f.53']).dt.days
    df_unprocessed.rename(columns={'f.20116': 'CAT_Smoking', 'f.21001': 'BMI', 'f.53': 'BaselineDate'}, inplace=True)

    df_unprocessed = df_unprocessed[df_unprocessed['CAT_Smoking'] != -3].copy()

    df_unprocessed = df_unprocessed.sort_values(by='date', na_position='last')
    df_unprocessed = df_unprocessed.reset_index()
    return df_unprocessed[~df_unprocessed.duplicated(subset=['summary', 'eid'], keep='first')].sort_values(by='eid').set_index('eid')

# %% Support functions to make the script modular (and to make it compatible with sex stratified analyses)

def load_and_split_data(df, covar=None):
    if bool(covar):
        df = df.join(covar, how='inner')
    male_data = df[df['Sex'] == 1]
    female_data = df[df['Sex'] == 0]
    return df, male_data, female_data

def logOR(data_test, col_y, pred_probs_test):
    try:
        data_test['predicted_prob'] = pred_probs_test
        data_test['quantile'] = pd.qcut(data_test['predicted_prob'], 5, labels=False, duplicates='drop') + 1
        odds_q3 = (data_test[data_test['quantile'] == 3][col_y].sum()) / \
                (data_test[data_test['quantile'] == 3][col_y].count() - data_test[data_test['quantile'] == 3][col_y].sum())
        odds_q5 = (data_test[data_test['quantile'] == 5][col_y].sum()) / \
                (data_test[data_test['quantile'] == 5][col_y].count() - data_test[data_test['quantile'] == 5][col_y].sum())
        return np.log(odds_q5 / odds_q3)
    except:
        return np.nan

def getF1(label, pred_prob, method="at50"):
    match method:
        case "otsu":        
            threshold = threshold_otsu(pred_prob.values)
        case "at50":
            threshold = 0.5
        case _:
            sys.exit(f"Method {method} for getF1 not implemented")
    pred = pred_prob > threshold
    return f1_score(label, pred)
    


@plots_and_results_by_sex
def get_scores(resDF, fold, method, res_type, diseaseDF, IIDs, sex, dis_col="BinCAT_Disease"):
    df = diseaseDF.join(resDF, how='inner')
    pred_col = resDF.name if type(resDF) is pd.Series else resDF.columns[0]
    df = df[df.index.isin(IIDs)]
    auc = roc_auc_score(df[dis_col], df[pred_col])
    logOR_value = logOR(df, pred_col, df[pred_col])
    f1_test = getF1(df[dis_col], df[pred_col], method="at50")
    return {"AUC_test": auc, "logOR_test": logOR_value, "F1_test": f1_test, "Sex": sex, "fold": fold, "method": method, "res_type": res_type}

@plots_and_results_by_sex
def get_scores_pancohort(resDF, fold, method, res_type, IIDs, sex):
    if type(resDF) is pd.Series: 
        pred_col = resDF.name
        df = pd.DataFrame(resDF[resDF.index.isin(IIDs)].rename(pred_col)) #the rename call is required for older pandas versions (e.g. 1.4.4)
    else:
        df = resDF[resDF.index.isin(IIDs)]
        pred_col = resDF.columns[0] 
    logOR_value = logOR(df, pred_col, df[pred_col])
    return {"logOR_test_pancohort": logOR_value, "Sex": sex, "fold": fold, "method": method, "res_type": res_type}

# %%
if __name__ == "__main__":
    parser = getARGSParser()
    args, _ = parser.parse_known_args()

    args.colour_box = args.colour_box.split(",") if bool(args.colour_box) else colours
    args.colour_prevalence = args.colour_prevalence.split(",") if bool(args.colour_prevalence) else colours
    args.colour_displots = args.colour_displots.split(",") if bool(args.colour_displots) else colours

    covars = pd.read_table(args.ext_covar, low_memory=False)
    IIDs_male = covars[covars.Sex==1].IID.to_list()
    IIDs_female = covars[covars.Sex==0].IID.to_list()

    # %% define the interesting comparisons [TODO: make it configurable. Currently, it's a "comment-out"-based approach]
    comparisons = [
        ('GLM', 'covar', 'Baseline', 1),
        ('Lasso', 'covar', 'Lasso Baseline', 0),

        ('GLM', 'covarNorm', 'Baseline (Normalised)', 0),
        ('Lasso', 'covarNorm', 'Lasso Baseline (Normalised)', 0),

        # ('GLM', 'nonPCCovar', 'non-PC Baseline', 0),
        # ('Lasso', 'nonPCCovar', 'Lasso non-PC Baseline', 0),

        ('GLM', 'singlePRS', 'max(Single PRS)', 0),

        ('GLM', 'singlePRSCovar', 'max(Single PRS + Baseline)', 1),
        ('Lasso', 'singlePRSCovar', 'max(Lasso(Single PRS + Baseline))', 0),
        
        ('GLM', 'singlePRSCovarNorm', 'max(norm(Single PRS + Baseline))', 0),
        ('Lasso', 'singlePRSCovarNorm', 'max(Lasso(norm(Single PRS + Baseline)))', 0),

        ('GLM', 'multiPRS', 'Multi PRS (GLM)', 0),
        ('Lasso', 'multiPRS', 'Multi PRS', 0),
        ('GLM', 'multiPRSCovar', 'Multi PRS + Baseline (GLM)', 0),
        ('Lasso', 'multiPRSCovar', 'Multi PRS + Baseline', 1),
        # ('GLM', 'multiPRSnonPCCovar', 'Multi PRS + non-PC Baseline (GLM)', 0),
        # ('Lasso', 'multiPRSnonPCCovar', 'Multi PRS + non-PC Baseline', 0),

        ('GLM', 'multiPRSNorm', 'Normalised Multi PRS (GLM)', 0),
        ('Lasso', 'multiPRSNorm', 'Normalised Multi PRS', 0),
        ('GLM', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (GLM)', 0),
        ('Lasso', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline', 0),
        # ('GLM', 'multiPRSNormnonPCCovar', 'Normalised Multi PRS + non-PC Baseline (GLM)', 0),
        # ('Lasso', 'multiPRSNormnonPCCovar', 'Normalised Multi PRS + non-PC Baseline', 0),

        ('LassoSteps0', 'multiPRSCovar', 'Multi PRS + Baseline (LassoSteps0)', 0),
        ('LassoSteps0', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (LassoSteps0)', 0),

        ('LassoSteps1', 'multiPRSCovar', 'Multi PRS + Baseline (LassoSteps1)', 0),
        ('LassoSteps1', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (LassoSteps1)', 0),

        ('LassoSteps2', 'multiPRSCovar', 'Multi PRS + Baseline (LassoSteps2)', 0),
        ('LassoSteps2', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (LassoSteps2)', 0),

        ('LassoSteps3', 'multiPRSCovar', 'Multi PRS + Baseline (LassoSteps3)', 0),
        ('LassoSteps3', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline (LassoSteps3)', 0),

        # ('ElasticNet', 'multiPRS', 'Multi PRS', 0),
        # ('ElasticNet', 'multiPRSCovar', 'Multi PRS + Baseline', 0),
        # # ('ElasticNet', 'multiPRSnonPCCovar', 'Multi PRS + non-PC Baseline', 0),

        # ('ElasticNet', 'multiPRSNorm', 'Normalised Multi PRS', 0),
        # ('ElasticNet', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline', 0),
        # # ('ElasticNet', 'multiPRSNormnonPCCovar', 'Normalised Multi PRS + non-PC Baseline', 0),

        # ('XGBoost', 'multiPRS', 'Multi PRS', 0),
        # ('XGBoost', 'multiPRSCovar', 'Multi PRS + Baseline', 0),
        # ('ElasticNet', 'multiPRSnonPCCovar', 'Multi PRS + non-PC Baseline', 0),

        # ('XGBoost', 'multiPRSNorm', 'Normalised Multi PRS', 0),
        # ('XGBoost', 'multiPRSNormCovar', 'Normalised Multi PRS + Baseline', 0),
        # ('ElasticNet', 'multiPRSNormnonPCCovar', 'Normalised Multi PRS + non-PC Baseline', 0),
    ]

    all_combinations = list(itertools.combinations(comparisons, 2))
    interesting_combinations = {(method, res_type) for method, res_type, _, _ in comparisons}
    tag_lookup = {(method, res_type): tag for method, res_type, tag, _ in comparisons}
    tag_order = [tag for _, _, tag, _ in comparisons]

    # %% obtain the results' pickles
    results = glob(f"{args.output_root}/*_raw_results.pkl")    


    # %% create summary of the scores and perform statistical tests

    if args.obtain_summary:
        collect_df = []
        collect_summary_df = []
        collect_stats = []

        with open(f"{args.output_root}/sumscores.txt", 'w') as f:
            for r in tqdm(results):
                d = os.path.basename(r).replace('_raw_results.pkl', '')
                f.write(f"\n\n{d}:-------------------------------\n")

                disease = pd.read_csv(f'{args.disease_root}/{d}.csv', low_memory=False, index_col="IID")
                disease["BinCAT_Disease"] = disease["BinCAT_Disease"].astype(int)

                with open(r, "rb") as pkl:
                    res = pickle.load(pkl)

                scores = []
                scores_pancohort = []
                for fold in res.keys():
                    for method in res[fold].keys():
                        if method == 'IDs':
                            continue
                        for res_type in res[fold][method].keys():
                            best_auc = {"Both": 0.0, "Female": 0.0, "Male": 0.0}
                            best_scores = {}
                            if "single" in res_type.lower():
                                for latent in res[fold][method][res_type]['pred_probs_test'].keys():
                                    scr = get_scores(resDF=res[fold][method][res_type]['pred_probs_test'][latent], diseaseDF=disease, IIDs_female=IIDs_female, IIDs_male=IIDs_male, fold=fold, method=method, res_type=res_type)
                                    for s in scr:
                                        if s['AUC_test'] > best_auc[s["Sex"]]:
                                            best_auc[s["Sex"]] = s['AUC_test']
                                            best_scores[s["Sex"]] = s
                                            best_scores[s["Sex"]]["Latent"] = latent
                                scores.append(list(best_scores.values()))
                                scores_pancohort.append(get_scores_pancohort(resDF=res[fold][method][res_type]['pred_probs_test_pancohort'][best_scores["Both"]["Latent"]], IIDs_female=IIDs_female, IIDs_male=IIDs_male, fold=fold, method=method, res_type=res_type))
                            else:
                                scores.append(get_scores(resDF=res[fold][method][res_type]['pred_probs_test'], diseaseDF=disease, IIDs_female=IIDs_female, IIDs_male=IIDs_male, fold=fold, method=method, res_type=res_type))
                                scores_pancohort.append(get_scores_pancohort(resDF=res[fold][method][res_type]['pred_probs_test_pancohort'], IIDs_female=IIDs_female, IIDs_male=IIDs_male, fold=fold, method=method, res_type=res_type))     
                scores = pd.DataFrame([score for sublist in scores for score in sublist])
                scores_pancohort = pd.DataFrame([score for sublist in scores_pancohort for score in sublist])
                df = pd.merge(scores, scores_pancohort, on=["Sex", "fold", "method", "res_type"])

                df['Disease'] = d.replace("-", " ")
                df['nPatients'] = len(disease) // 2

                df = df[df.apply(lambda row: (row['method'], row['res_type']) in interesting_combinations, axis=1)]
                df['tag'] = df.apply(lambda row: tag_lookup.get((row['method'], row['res_type'])), axis=1)
                df['tag'] = pd.Categorical(df['tag'], categories=tag_order, ordered=True)
                collect_df.append(df)

                summary_df = getSummaryDF(df, pancohort=args.is_pancohort)
                f.write("\nSummary scores:-")
                f.write(summary_df.to_string(index=False))
                summary_df['Disease'] = d.replace("-", " ")
                summary_df['nPatients'] = len(disease) // 2
                collect_summary_df.append(summary_df)

                summary_stats = compareSummaryPairs(df=df, all_combinations=all_combinations, disease=d.replace("-", " "), pancohort=args.is_pancohort)
                summary_stats = pd.concat(summary_stats)
                
                auc_delong = compareAUCPairs(res=res, all_combinations=all_combinations, disease=d.replace("-", " "), diseaseDF=disease, IIDs_female=IIDs_female, IIDs_male=IIDs_female, resDF=df)
                auc_delong = pd.concat(auc_delong)

                stats_combined = pd.merge(auc_delong, summary_stats, on=["Disease", "Comparison", "Sex", "Method0", "Method1"]) #Method0 and Method1, along with Comparison, are redundant. But, I'm keeping them just to avoid those columns appearing twice with _x and _y suffixes. 
                stats_combined['nPatients'] = len(disease) // 2
                collect_stats.append(stats_combined)

                stats2print = stats_combined[['Comparison', 'Sex', 'p_value_DeLong_AUC']].copy()
                stats2print['is Sig'] = stats2print['p_value_DeLong_AUC'].apply(lambda x: 'Sig' if x < 0.05 else 'Not Sig')
                f.write("\n\nDeLong's p-values:-")
                f.write(stats2print.to_string(index=False))

                f.write("\n\n-----------------------------------\n\n")

        collect_df = pd.concat(collect_df)
        collect_summary_df = pd.concat(collect_summary_df)
        collect_stats = pd.concat(collect_stats)

        collect_df.to_csv(f"{args.output_root}/combined_results.tsv", sep="\t", index=False)
        collect_summary_df.to_csv(f"{args.output_root}/combined_summary_results.tsv", sep="\t", index=False)
        collect_stats.to_csv(f"{args.output_root}/combined_stats_sig.tsv", sep="\t", index=False)

    else:
        collect_df = pd.read_table(f"{args.output_root}/combined_results.tsv", low_memory=False)
        collect_summary_df = pd.read_table(f"{args.output_root}/combined_summary_results.tsv", low_memory=False)
        collect_stats = pd.read_table(f"{args.output_root}/combined_stats_sig.tsv", low_memory=False)

    # %%
    if bool(args.drop_dis):
        print(f"Dropping {args.drop_dis} from the analysis")
        collect_df = collect_df[~collect_df.Disease.isin(args.drop_dis.split(","))]
        collect_summary_df = collect_summary_df[~collect_summary_df.Disease.isin(args.drop_dis.split(","))]
        collect_stats = collect_stats[~collect_stats.Disease.isin(args.drop_dis.split(","))]

    if args.dis_plots_rawPRS:
        collect_df_rawPRS = collect_df[collect_df.tag == "max(Single PRS)"]

    if args.drop_comparisons:
        selected_comparisons = []
        for c in comparisons:
            if c[3] == 1:
                selected_comparisons.append(c)
        interesting_combinations = {(method, res_type) for method, res_type, _, _ in selected_comparisons}
        tag_lookup = {(method, res_type): tag for method, res_type, tag, _ in selected_comparisons}
        tag_order = [tag for _, _, tag, _ in selected_comparisons]
        collect_df = collect_df[collect_df.tag.isin(tag_order)]
        collect_summary_df = collect_summary_df[collect_summary_df.tag.isin(tag_order)]
        collect_stats = collect_stats[collect_stats.Method0.isin(tag_order) & collect_stats.Method1.isin(tag_order)]  
        comparisons = selected_comparisons

    # %% print the pairwise improvements

    if args.obtain_pairwise_improvements:    
        combined_df = combineDFs(collect_stats, collect_summary_df)
        tab_summary = summary_table(combined_df)
        for metric in ["AUC", "logOR", "logOR_Pancohort", "F1"]:
            create_excel_summary(tab_summary, metric, f"{args.output_root}/summary_{metric}.xlsx")
        combined_df.to_csv(f"{args.output_root}/summary_combined.tsv", sep="\t", index=False)
        tab_summary.to_csv(f"{args.output_root}/summary_table.tsv", sep="\t", index=False)

    # %% "improve" the dataframes for plotting
    collect_df.Disease = collect_df.Disease.apply(lambda x: transform_text(x, "sentence_case"))
    if args.dis_plots_rawPRS:
        collect_df_rawPRS.Disease = collect_df_rawPRS.Disease.apply(lambda x: transform_text(x, "sentence_case"))

    # %% sort the diseases based on the AUC
    sorted_diseases = collect_df[(collect_df['res_type'] == 'singlePRSCovar') & (collect_df.Sex == 'Both')].groupby('Disease')['AUC_test'].mean().sort_values(ascending=False).index
    match (args.sort_mode_top_dis):
        case "0":
            pass
        case "1":
            sorted_diseases = sorted(sorted_diseases)
        case _ :
            print("Using the provided custom order...")
            sorted_diseases = args.sort_mode_top_dis.split(",")

    if args.split_top_dis_half:
        top_diseases = sorted_diseases[:(len(sorted_diseases)//2)]
        bottom_diseases = sorted_diseases[(len(sorted_diseases)//2):]
    else:
        top_diseases = sorted_diseases
        bottom_diseases = []

    df_top = collect_df[collect_df['Disease'].isin(top_diseases)].copy()
    df_top['Disease'] = pd.Categorical(df_top['Disease'], categories=top_diseases, ordered=True)

    if bool(bottom_diseases):
        df_bottom = collect_df[collect_df['Disease'].isin(bottom_diseases)].copy()
        df_bottom['Disease'] = pd.Categorical(df_bottom['Disease'], categories=bottom_diseases, ordered=True)

    if args.save_raw:
        os.makedirs(os.path.join(args.output_root, "plots", "raw_store"), exist_ok=True)

    # %% Create box plots for AUC_test, top N//2 diseases and bottom N//2 diseases seperately

    if args.plot_box_AUC:
        os.makedirs(os.path.join(args.output_root, "plots", "auc"), exist_ok=True)

        plot_box(df_top, x='Disease', y='AUC_test', y_tag='AUC', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/auc/top_{len(top_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_AUC)
        if bool(bottom_diseases):
            plot_box(df_bottom, x='Disease', y='AUC_test', y_tag='AUC', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/auc/bottom_{len(bottom_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_AUC)
    
    # %% Create box plots for F1_test, top N//2 diseases and bottom N//2 diseases seperately

    if args.plot_box_F1:
        os.makedirs(os.path.join(args.output_root, "plots", "f1"), exist_ok=True)

        plot_box(df_top, x='Disease', y='F1_test', y_tag='F1', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/f1/top_{len(top_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_F1)
        if bool(bottom_diseases):
            plot_box(df_bottom, x='Disease', y='F1_test', y_tag='F1', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/f1/bottom_{len(bottom_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_F1)
    
    # %% Create box plots for logOR_test, top N//2 diseases and bottom N//2 diseases seperately

    if args.plot_box_logOR:
        os.makedirs(os.path.join(args.output_root, "plots", "logOR"), exist_ok=True)

        plot_box(df_top, x='Disease', y='logOR_test', y_tag='logOR', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/logOR/top_{len(top_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_logOR)
        if bool(bottom_diseases):
            plot_box(df_bottom, x='Disease', y='logOR_test', y_tag='logOR', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/logOR/bottom_{len(bottom_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_logOR)
    
    # %% Create box plots for logOR_test_pancohort, top N//2 diseases and bottom N//2 diseases seperately
    if args.is_pancohort and args.plot_box_logOR:
        os.makedirs(os.path.join(args.output_root, "plots", "logOR_pancohort"), exist_ok=True)

        plot_box(df_top, x='Disease', y='logOR_test_pancohort', y_tag='logOR (Pancohort)', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/logOR_pancohort/top_{len(top_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_logOR_pancohort)
        if bool(bottom_diseases):
            plot_box(df_bottom, x='Disease', y='logOR_test_pancohort', y_tag='logOR (Pancohort)', legends_tag="Method", hue='tag', hue_order=tag_order, palette=args.colour_box, colour_transparency=args.colour_transparency_box, save_path_noext=f"{args.output_root}/plots/logOR_pancohort/bottom_{len(bottom_diseases)}_diseases" if args.save_plots else "", save_format=args.save_format, y_tick_limits=args.box_limits_logOR_pancohort)

    # %% Create prevalence vs PRS scores plots for the pancohort, for the best single latent PRS

    if args.plot_prevalence_prs_pancohort:

        if args.plot_prevalence_prs_pancohort:
            os.makedirs(os.path.join(args.output_root, "plots", "prevalence_prs_pancohort"), exist_ok=True)

        raw_res_store = []
        for dis in sorted_diseases:
            r = f"{args.output_root}/{dis.replace(' ', '-').lower()}_raw_results.pkl"

            with open(r, "rb") as pkl:
                res = pickle.load(pkl)
        
            best_AUC = 0
            best_prs_latent = None
            for fold in res.keys():
                _, res_method = getIndividualProbs(res[fold]['GLM']['singlePRS'], 'singlePRS', IIDs=IIDs_female+IIDs_male) #Maybe we can think about also doing it in a sex sex-stratified manner?
                if res_method['AUC_test'] > best_AUC:
                    best_AUC = res_method['AUC_test']
                    best_prs_latent = res_method.name.split("PRS:")[-1]
                
            pth_prs = f"{args.rds_pres_prefix}{best_prs_latent}{args.rds_pres_suffix}"

            data = pyreadr.read_r(pth_prs.replace(args.rds_tag_prs, args.tag_data))[None]
            prs = data[['IID']]
            r_obj = read_rds(pth_prs)
            prs["PRS"] = r_obj['data'][r_obj['attributes']['names']['data'].index(args.tag_prs)]['data']
            prs.set_index("IID", inplace=True)

            d = os.path.basename(r).replace('_raw_results.pkl', '')
            disease = pd.read_csv(f'{args.disease_root}/{d}.csv', low_memory=False, index_col="IID")
            combined = prs.join(disease[['BinCAT_Disease']])
            combined.fillna(0, inplace=True)
            combined["BinCAT_Disease"] = combined["BinCAT_Disease"].astype(int)

            if args.save_plots:
                pth = f"{args.output_root}/plots/prevalence_prs_pancohort/{d.replace(' ', '-')}.{args.save_format}"
            else:
                pth = ""

            d = d.replace("-", " ")
            if args.sex_stratified:
                data.IID = data.IID.astype(combined.index.dtype)
                IIDs_all = list(combined.index)
                IIDs_female = list(set(IIDs_all).intersection(set(list(data[data.Sex == "0"].IID))))
                IIDs_male = list(set(IIDs_all).intersection(set(list(data[data.Sex == "1"].IID))))
                combined.loc[IIDs_female, 'PRS_Female'] = combined.loc[IIDs_female, 'PRS']
                combined.loc[IIDs_male, 'PRS_Male'] = combined.loc[IIDs_male, 'PRS']
                pth = pth.replace(f".{args.save_format}", f"_sex_stratified.{args.save_format}")
                raw_res_store.append(plot_prevalence(combined, res_keys=['PRS_Female', 'PRS', 'PRS_Male'], res_labels=['Female', 'Both', 'Male'], res_colours=args.colour_prevalence[:3], res_fmts=markers[:3], col_disease='BinCAT_Disease', plot_title=f'max(Single PRS) Risk scores: {d}', individual_prev_line=True, save_path=pth, save_raw=args.save_raw))
            else:
                raw_res_store.append(plot_prevalence(combined, res_keys=['PRS'], res_labels=['max(Single PRS)'], res_colours=['blue'], res_fmts=['o'], col_disease='BinCAT_Disease', plot_title=f'Risk scores: {d}', save_path=pth, save_raw=args.save_raw))

        if args.save_raw:
            with open(f'{os.path.join(args.output_root, "plots", "raw_store")}/raw_prevalence_prs_pancohort.pkl', "wb") as pkl:
                pickle.dump(raw_res_store, pkl)

    # %% Create prevalence vs probability plots for the pancohort, for the best fold
    
    raw_res_store = []
    if args.is_pancohort and args.plot_prevalence_prob_pancohort:
        os.makedirs(os.path.join(args.output_root, "plots", "prevalence_prob_pancohort"), exist_ok=True)

        for dis in sorted_diseases:
            r = f"{args.output_root}/{dis.replace(' ', '-').lower()}_raw_results.pkl"

            with open(r, "rb") as pkl:
                res = pickle.load(pkl)

            d = os.path.basename(r).replace('_raw_results.pkl', '')
            diseaseDF = pd.read_csv(f'{args.disease_root}/{d}.csv', low_memory=False, index_col="IID")

            raw_res_store.append(plot_prevalence_prob_pancohort(res, dis, diseaseDF, collect_df, IIDs_female=IIDs_female, IIDs_male=IIDs_male, comparisons=comparisons, save_format=args.save_format))

        if args.save_raw:
            with open(f'{os.path.join(args.output_root, "plots", "raw_store")}/raw_prevalence_prob_pancohort.pkl', "wb") as pkl:
                pickle.dump(raw_res_store, pkl)
        
    # %% Create prevalence vs probability plots for the pancohort, for the best fold

    if args.plot_cum_disease_burden or args.plot_cum_hazard or args.plot_KM_survival:

        if args.plot_cum_disease_burden:
            os.makedirs(os.path.join(args.output_root, "plots", "cumulative_disease_burden"), exist_ok=True)
            if args.dis_plots_rawPRS:
                os.makedirs(os.path.join(args.output_root, "plots", "rawPRS_cumulative_disease_burden"), exist_ok=True)
        if args.plot_cum_hazard:
            os.makedirs(os.path.join(args.output_root, "plots", "cumulative_hazard"), exist_ok=True)
            if args.dis_plots_rawPRS:
                os.makedirs(os.path.join(args.output_root, "plots", "rawPRS_cumulative_hazard"), exist_ok=True)
        if args.plot_KM_survival:
            os.makedirs(os.path.join(args.output_root, "plots", "KM_survival"), exist_ok=True)
            if args.dis_plots_rawPRS:
                os.makedirs(os.path.join(args.output_root, "plots", "rawPRS_KM_survival"), exist_ok=True)

        if args.reprocess_raw_diseases or not os.path.isfile(f"{args.output_root}/raw_diseases.tsv"):
            raw_diseases = get_disease_info(disease_path=args.raw_disease_path, raw_baseline_path=args.raw_baseline_path, info_assessment_centre_path=args.raw_centre_info_path, cutoff_date=args.dis_plots_cutoff_date)
            raw_diseases = raw_diseases.join(covars.set_index("IID")['Sex'], how='inner')
            raw_diseases = add_grouped_diseases(raw_diseases, cardiac_mappings, "healthy")
            raw_diseases.to_csv(f"{args.output_root}/raw_diseases.tsv", sep="\t")
        else:
            print("Existing raw_diseases.tsv found! Loading...")
            raw_diseases = pd.read_table(f"{args.output_root}/raw_diseases.tsv", low_memory=False).set_index("eid")
            raw_diseases = add_grouped_diseases(raw_diseases, cardiac_mappings, "healthy") #TODO: remove it
        
        if args.save_raw:
            os.makedirs(os.path.join(args.output_root, "plots", "raw_displots"), exist_ok=True)
            raw_res_store = []

        for dis in sorted_diseases:
            r = f"{args.output_root}/{dis.replace(' ', '-').lower()}_raw_results.pkl"

            with open(r, "rb") as pkl:
                res = pickle.load(pkl)
            
            if bool(args.dis_plots_mod):
                raw_res_store.append(plot_dis_plots(res=res, disease=dis, diseaseDF=raw_diseases, collect_df=collect_df, args=args, IIDs_female=IIDs_female, IIDs_male=IIDs_male, model_type=args.dis_plots_mod.split(',')[0], model=args.dis_plots_mod.split(',')[1], PRS_col="predicted", ending_year = args.dis_plots_upto_Nyear, use_raw_PRS=False, res_colours=args.colour_displots, save_format=args.save_format, save_raw=args.save_raw))
            
                if args.dis_plots_rawPRS:
                    raw_res_store.append(plot_dis_plots(res=res, disease=dis, diseaseDF=raw_diseases, collect_df=collect_df_rawPRS, args=args, IIDs_female=IIDs_female, IIDs_male=IIDs_male, model_type="GLM", model="singlePRS", PRS_col="predicted", ending_year = args.dis_plots_upto_Nyear, use_raw_PRS=True, res_colours=args.colour_displots, save_format=args.save_format, save_raw=args.save_raw))

        if args.save_raw:
            with open(f'{os.path.join(args.output_root, "plots", "raw_store")}/raw_displots.pkl', "wb") as pkl:
                pickle.dump(raw_res_store, pkl)