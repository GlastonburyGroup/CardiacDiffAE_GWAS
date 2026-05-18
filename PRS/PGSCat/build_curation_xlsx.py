#!/usr/bin/env python
"""build_curation_xlsx.py

Populate the **PGS Catalog Curation Template v11**
(``PGSSubmission_CurationTemplate_v11.xlsx``, see
``PGSCatalog_CurationGuidelines_v5.pdf``) with:

* one row in **Publication Information** (placeholder bibliographic info),
* one row per latent in the **Score(s)** sheet,
* two rows per latent in **Sample Descriptions** (UKB training + UKB
  testing cohort skeletons),
* an empty **Performance Metrics** sheet that the user fills manually.

The template uses a two-row header on the Score(s) and Performance
Metrics sheets (row 1 = section banner, row 2 = column names, data
starts at row 3). Sample Descriptions has a single header row (row 1)
and data starts at row 2.

Layouts below are hardcoded to v11; if the template is updated the
hardcoded mappings here must be revisited.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


# -------------------- v11 layout (1-indexed) --------------------

PUB_SHEET = "Publication Information"
PUB_DATA_ROW = 2  # row 1 is header
PUB_COLS = {
    "pubmed_id":          1,  # A
    "doi":                2,  # B
    "journal":            3,  # C
    "publication_date":   4,  # D
    "first_author_last":  5,  # E
    "first_author_init":  6,  # F
}

SCORES_SHEET = "Score(s)"
SCORES_HEADER_ROW = 2
SCORES_DATA_START = 3
SCORES_COLS = {
    "score_name":      1,   # A
    "reported_trait":  2,   # B
    "trait_info":      3,   # C
    "efo_ids":         4,   # D
    "efo_names":       5,   # E
    "method_name":     6,   # F
    "method_details":  7,   # G
    "genome_build":    8,   # H
    "variants_number": 9,   # I
    "interactions":    10,  # J
    "curation_notes":  11,  # K
}

SAMPLES_SHEET = "Sample Descriptions"
SAMPLES_DATA_START = 2
SAMPLES_COLS = {
    "associated_scores":   1,   # A
    "study_stage":         2,   # B
    "sample_set_id":       3,   # C
    "gwas_catalog_id":     4,   # D
    "pmid_doi":            5,   # E
    "n_individuals":       6,   # F
    "n_cases":             7,   # G
    "n_controls":          8,   # H
    "pct_male":            9,   # I
    "age":                 10,  # J
    "broad_ancestry":      11,  # K
    "ancestry":            12,  # L
    "country":             13,  # M
    "ancestry_extra":      14,  # N
    "phenotype_def":       15,  # O
    "followup_time":       16,  # P
    "cohorts":             17,  # Q
    "additional_info":     18,  # R
    "curation_notes":      19,  # S
}

# Study stage values from the Sample Refr. sheet:
STAGE_GWAS = "GWAS/Variant associations"
STAGE_DEVELOPMENT = "Score development"
STAGE_TESTING = "Testing"


def write_publication(wb, args):
    ws = wb[PUB_SHEET]
    r = PUB_DATA_ROW
    if args.pubmed_id:
        ws.cell(row=r, column=PUB_COLS["pubmed_id"], value=args.pubmed_id)
    if args.doi:
        ws.cell(row=r, column=PUB_COLS["doi"], value=args.doi)
    if args.journal:
        ws.cell(row=r, column=PUB_COLS["journal"], value=args.journal)
    if args.pub_date:
        ws.cell(row=r, column=PUB_COLS["publication_date"], value=args.pub_date)
    if args.author_last:
        ws.cell(row=r, column=PUB_COLS["first_author_last"], value=args.author_last)
    if args.author_init:
        ws.cell(row=r, column=PUB_COLS["first_author_init"], value=args.author_init)
    print(f"[INFO] Publication Information: row {r} filled")


def write_scores(wb, args, latents, combined=None):
    """Write one row per latent (and optional combined per-disease scores)
    into the Score(s) sheet."""
    ws = wb[SCORES_SHEET]
    cols = SCORES_COLS
    r = SCORES_DATA_START
    for latent, n_variants in latents:
        ws.cell(row=r, column=cols["score_name"], value=latent)
        ws.cell(row=r, column=cols["reported_trait"],
                value=f"{args.trait_reported_prefix}{latent}")
        ws.cell(row=r, column=cols["trait_info"], value=args.trait_info)
        # Default EFO mapping for cardiac MRI imaging-derived latent features:
        # EFO:0022611 = "magnetic resonance imaging of the heart"
        # (https://www.ebi.ac.uk/efo/EFO_0022611). May be overridden via
        # --trait_efo / --trait_efo_name.
        if args.trait_efo:
            ws.cell(row=r, column=cols["efo_ids"], value=args.trait_efo)
        if args.trait_efo_name:
            ws.cell(row=r, column=cols["efo_names"], value=args.trait_efo_name)
        ws.cell(row=r, column=cols["method_name"], value=args.method_name)
        ws.cell(row=r, column=cols["method_details"], value=args.method_details)
        ws.cell(row=r, column=cols["genome_build"], value=args.genome_build)
        ws.cell(row=r, column=cols["variants_number"], value=int(n_variants))
        ws.cell(row=r, column=cols["interactions"], value=0)
        ws.cell(row=r, column=cols["curation_notes"],
                value=(f"Scoring file: {latent}.txt.gz. "
                       "Derived from LDpred2-auto consensus betas "
                       "(see method details)."))
        r += 1
    n_latent_rows = r - SCORES_DATA_START

    # ---- Combined per-disease scores (back-calculated SNP-level weights) ----
    if combined:
        for disease, n_variants in combined:
            ws.cell(row=r, column=cols["score_name"], value=disease)
            ws.cell(row=r, column=cols["reported_trait"], value=disease)
            ws.cell(row=r, column=cols["trait_info"],
                    value=("Multi-PRS combined risk score: linear combination "
                           "of 49 per-latent LDpred2-auto PGS using "
                           "disease-specific LASSO coefficients. Covariate "
                           "terms are reported separately (sidecar TSV)."))
            # EFO IDs / Names are disease-specific - left blank for manual curation.
            ws.cell(row=r, column=cols["method_name"],
                    value="LDpred2-auto + LASSO multi-PRS combination")
            ws.cell(row=r, column=cols["method_details"], value=args.combined_method_details)
            ws.cell(row=r, column=cols["genome_build"], value=args.genome_build)
            ws.cell(row=r, column=cols["variants_number"], value=int(n_variants))
            ws.cell(row=r, column=cols["interactions"], value=0)
            ws.cell(row=r, column=cols["curation_notes"],
                    value=(f"Scoring file: {disease}.txt.gz. "
                           f"Covariate sidecar: {disease}_covariate_coefficients.tsv "
                           "(not folded into per-SNP weights)."))
            r += 1
    n_combined_rows = (r - SCORES_DATA_START) - n_latent_rows
    print(f"[INFO] Score(s): wrote {n_latent_rows} latent rows + "
          f"{n_combined_rows} combined-disease rows (rows "
          f"{SCORES_DATA_START}..{r - 1})")


def write_samples(wb, args, latent_ids):
    """Write paired training (Score development) + testing rows per latent.

    For UKBB-derived PGS the training and testing cohorts are the same
    UKB cohort (subdivided across CV folds), but PGS Catalog requires
    them on separate rows with separate Study Stage values.
    """
    ws = wb[SAMPLES_SHEET]
    cols = SAMPLES_COLS
    r = SAMPLES_DATA_START
    all_scores = ", ".join(latent_ids)

    # ---- Variant associations (GWAS) - per UKB regenie run ----
    ws.cell(row=r, column=cols["associated_scores"], value=all_scores)
    ws.cell(row=r, column=cols["study_stage"], value=STAGE_GWAS)
    ws.cell(row=r, column=cols["pmid_doi"], value=args.gwas_pmid_doi)
    ws.cell(row=r, column=cols["n_individuals"], value=args.gwas_n)
    ws.cell(row=r, column=cols["broad_ancestry"], value=args.broad_ancestry)
    ws.cell(row=r, column=cols["ancestry"], value=args.ancestry)
    ws.cell(row=r, column=cols["country"], value=args.country)
    ws.cell(row=r, column=cols["phenotype_def"], value=args.gwas_phenotype_def)
    ws.cell(row=r, column=cols["cohorts"], value=args.cohort_id)
    ws.cell(row=r, column=cols["additional_info"], value=args.gwas_additional)
    r += 1

    # ---- Score development (LDpred2-auto + Lasso training) ----
    ws.cell(row=r, column=cols["associated_scores"], value=all_scores)
    ws.cell(row=r, column=cols["study_stage"], value=STAGE_DEVELOPMENT)
    ws.cell(row=r, column=cols["n_individuals"], value=args.train_n)
    ws.cell(row=r, column=cols["broad_ancestry"], value=args.broad_ancestry)
    ws.cell(row=r, column=cols["ancestry"], value=args.ancestry)
    ws.cell(row=r, column=cols["country"], value=args.country)
    ws.cell(row=r, column=cols["cohorts"], value=args.cohort_id)
    ws.cell(row=r, column=cols["additional_info"], value=args.train_additional)
    r += 1

    # ---- Testing (held-out UKB folds, used for performance metrics) ----
    ws.cell(row=r, column=cols["associated_scores"], value=all_scores)
    ws.cell(row=r, column=cols["study_stage"], value=STAGE_TESTING)
    ws.cell(row=r, column=cols["sample_set_id"], value=args.test_sample_set_id)
    ws.cell(row=r, column=cols["n_individuals"], value=args.test_n)
    ws.cell(row=r, column=cols["broad_ancestry"], value=args.broad_ancestry)
    ws.cell(row=r, column=cols["ancestry"], value=args.ancestry)
    ws.cell(row=r, column=cols["country"], value=args.country)
    ws.cell(row=r, column=cols["cohorts"], value=args.cohort_id)
    ws.cell(row=r, column=cols["additional_info"], value=args.test_additional)
    r += 1

    print(f"[INFO] Sample Descriptions: wrote 3 rows (GWAS / development / "
          f"testing); rows {SAMPLES_DATA_START}..{r - 1}")


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--template", required=True,
                   help="Path to PGSSubmission_CurationTemplate_v11.xlsx")
    p.add_argument("--raw_dir", required=True,
                   help="Directory of per-latent *.scoring.raw.tsv")
    p.add_argument("--combined_dir", default="",
                   help=("Optional directory of combined per-disease "
                         "*_combined.scoring.raw.tsv files (output of "
                         "derive_combined_pgs.py). When provided, one extra "
                         "row per disease is added to the Score(s) sheet."))
    p.add_argument("--out", required=True, help="Filled .xlsx output path")

    # Publication
    p.add_argument("--pubmed_id", default="")
    p.add_argument("--doi", default="")
    p.add_argument("--journal", default="")
    p.add_argument("--pub_date", default="",
                   help="dd-mm-yyyy (per template hint)")
    p.add_argument("--author_last", default="Chatterjee")
    p.add_argument("--author_init", default="S")

    # Score(s)
    p.add_argument("--trait_reported_prefix",
                   default="Cardiac MRI imaging-derived latent feature ")
    p.add_argument("--trait_info",
                   default="Latent feature from a diffusion autoencoder trained on "
                           "UK Biobank short-axis cardiac MRI (field 20208), "
                           "selected for downstream multi-PRS disease prediction.")
    p.add_argument("--trait_efo", default="EFO:0022611",
                   help=("EFO ID for Reported Trait. Default: EFO:0022611 "
                         "(magnetic resonance imaging of the heart)."))
    p.add_argument("--trait_efo_name",
                   default="magnetic resonance imaging of the heart",
                   help="EFO term name matching --trait_efo.")
    p.add_argument("--method_name", default="LDpred2-auto")
    p.add_argument("--method_details",
                   default=("LDpred2-auto from bigsnpr (Prive et al.); "
                            "vec_p_init=seq_log(1e-4,0.9,length.out=ncores); "
                            "h2_init=LDSC estimate; chains kept if "
                            "|sd(pred_i)-median(sd)|<3*MAD; "
                            "final_beta=rowMeans of retained beta_est. "
                            "LD reference: UKB unrelated Caucasian (KING<0.0625), "
                            "MAF>0.01, INFO>0.4, lw_gw conditionally-independent SNPs."))
    p.add_argument("--combined_method_details",
                   default=("Per-latent LDpred2-auto betas (49 latents) linearly "
                           "combined using disease-specific LASSO coefficients "
                           "aggregated across 5 CV folds. The LASSO model was fit "
                           "on raw-scale PRS values plus covariates (20 PCs, age, "
                           "sex). Covariate terms are NOT folded into per-SNP "
                           "weights and are distributed as a sidecar TSV."))
    p.add_argument("--genome_build", default="GRCh37")

    # Sample Descriptions
    p.add_argument("--broad_ancestry", default="European",
                   help="Broad Ancestral Category (from Ancestry Refr. sheet)")
    p.add_argument("--ancestry", default="British, Irish, Other White",
                   help="Detailed ancestry description")
    p.add_argument("--country", default="United Kingdom")
    p.add_argument("--cohort_id", default="UKB",
                   help="Comma-separated cohort short IDs (see Cohort Refr.)")
    p.add_argument("--gwas_pmid_doi", default="",
                   help="PMID/DOI of the GWAS source publication")
    p.add_argument("--gwas_n", type=int, default=0,
                   help="N participants in the underlying GWAS")
    p.add_argument("--gwas_phenotype_def",
                   default=("Quantitative residualised latent feature (rank-INT) "
                            "from cardiac MRI auto-encoder; GWAS run with REGENIE."))
    p.add_argument("--gwas_additional",
                   default="UK Biobank field 20208 (short-axis cine images).")
    p.add_argument("--train_n", type=int, default=0,
                   help="N participants used for LDpred2-auto training")
    p.add_argument("--train_additional",
                   default=("LDpred2-auto fit using UK Biobank Caucasian unrelated "
                            "subset (KING coefficient < 0.0625)."))
    p.add_argument("--test_sample_set_id", default="UKB_holdout_CV",
                   help="ID linking testing samples to Performance Metrics rows")
    p.add_argument("--test_n", type=int, default=0,
                   help="N participants used for evaluation (held-out folds)")
    p.add_argument("--test_additional",
                   default=("Evaluation of disease-prediction multi-PRS Lasso models "
                            "via 5-fold cross-validation on UK Biobank."))

    args = p.parse_args()

    raw_dir = Path(args.raw_dir)
    raw_files = sorted(raw_dir.glob("*.scoring.raw.tsv"))
    if not raw_files:
        print(f"[ERROR] no *.scoring.raw.tsv in {raw_dir}", file=sys.stderr)
        return 1

    latents = []
    for raw in raw_files:
        latent = raw.name.replace(".scoring.raw.tsv", "")
        with raw.open() as fh:
            n = sum(1 for _ in fh) - 1
        latents.append((latent, n))

    combined: list[tuple[str, int]] = []
    if args.combined_dir:
        cdir = Path(args.combined_dir)
        for raw in sorted(cdir.glob("*_combined.scoring.raw.tsv")):
            disease = raw.name.replace("_combined.scoring.raw.tsv", "")
            with raw.open() as fh:
                n = sum(1 for _ in fh) - 1
            combined.append((disease, n))
        print(f"[INFO] combined per-disease scores: {len(combined)} found in {cdir}")

    wb = load_workbook(args.template)
    for required in (PUB_SHEET, SCORES_SHEET, SAMPLES_SHEET):
        if required not in wb.sheetnames:
            print(f"[ERROR] template missing sheet: {required!r}. "
                  f"Sheets: {wb.sheetnames}", file=sys.stderr)
            return 2

    write_publication(wb, args)
    write_scores(wb, args, latents, combined=combined)
    write_samples(wb, args, [lat for lat, _ in latents])

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[OK] saved -> {out_path}")
    print()
    print("Manual completion still required in the saved workbook:")
    print("  * Publication Information row 2: confirm PMID / DOI when known.")
    print("  * Score(s) sheet: review trait wording per latent.")
    print("  * Sample Descriptions: fill N / age / sex / ancestry counts.")
    print("  * Performance Metrics sheet: add one row per (score,sample_set)")
    print("    with OR/AUROC/R^2 from the paper.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
